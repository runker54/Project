# coding:utf-8
import xlrd
import openpyxl
import os


WB = openpyxl.Workbook()
ws = WB.active
path = r'C:\Users\65680\Desktop\赫章县种植情况调查表收集20200608'
path_list = []
for roots, dirs, files in os.walk(path):
    for file in files:
        if file[-3:].lower() == 'xls':
            ck = os.path.join(roots, file)
            path_list.append(ck)

for _id, cee in enumerate(path_list):
    ws.cell(_id+1, 1, cee)
# r = 0
# for _d in path_list:
#     if _d[-3:].lower() == 'xls':
#         # data = os.path.join(path, _d)
#         print(f'{_d}')
#         work = xlrd.open_workbook(_d)
#         sheet = work.sheet_by_index(0)
#         rows = sheet.nrows
#         columns = sheet.ncols
#         for _r in range(5, rows-1):
#             for _c in range(columns):
#                 cell_v = sheet.cell_value(_r, _c)
#                 print(cell_v)
#                 ws.cell(_r-4+r, _c+1, cell_v)
#         r = r+rows
WB.save(r'C:\Users\65680\Desktop\目录.xlsx')




