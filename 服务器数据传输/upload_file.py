#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Date:2021-11-25
# Author:Runker54
# ----------------------
import paramiko
import openpyxl
import datetime
import time
import os
import random
hostname = '120.77.46.249'

username = 'root'

password = 'runker54.++'

port = 22  # 配置信息可以写到配置文件中
client = paramiko.SSHClient()
client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
client.connect(hostname=hostname, username=username, password=password)
trans = client.get_transport()
sftp = paramiko.SFTPClient.from_transport(trans)

# loacl_file是要上传的本地文件路径

# remote_path是要上传到服务器上指定文件的路径
dir_path = r"C:\Users\65680\Desktop\样品采集名单\地址路径表.xlsx"
wb_a = openpyxl.load_workbook(dir_path)
ws_a = wb_a.active
rows = ws_a.max_row
dict_adress = {}
for one_row in range(2,rows):
    bianma = ws_a[f"F{one_row}"].value
    leixin = ws_a[f"H{one_row}"].value
    url = ws_a[f"G{one_row}"].value
    dict_adress[bianma+leixin] =url

shuidao_dir = r"E:\pic_path\水稻地块_压缩后"
shuidao_pic_list = [os.path.join(shuidao_dir,_x) for _x in os.listdir(shuidao_dir)]
yumi_dir = r"E:\pic_path\玉米地块_压缩后"
yumi_pic_list = [os.path.join(yumi_dir,_x) for _x in os.listdir(yumi_dir)]
pinzhong_dict = {}
pinzhong_path = r"E:\pic_path\水稻品种"
for one_pinzhogn in os.listdir(pinzhong_path):
    name = one_pinzhogn.split(".")[0]
    pinzhong_dict[name] = os.path.join(pinzhong_path, one_pinzhogn)

def upload(local_file, remote_path):
    try:
        print('开始上传文件%s ' % datetime.datetime.now())
        sftp.put(local_file, remote_path)
        print("从本地： %s 上传到： %s" % (local_file, remote_path))
        print('文件上传成功 %s ' % datetime.datetime.now())
        # trans.close()
    except Exception as e:
        print(repr(e))
if __name__ == '__main__':
    data_path = r"C:\Users\65680\Desktop\update.xlsx"
    wb = openpyxl.load_workbook(data_path)
    ws = wb.active
    rows = ws.max_row
    # for one_row in range(2, rows+1):
    #     bianma = ws[f"A{one_row}"].value
    #     tudileixin = ws[f"B{one_row}"].value
    #     pinzhong = ws[f"D{one_row}"].value
    #     if tudileixin == "水田":
    #         bianma_pic = shuidao_pic_list[random.randint(0, len(shuidao_pic_list))-1]
    #         up_dic = dict_adress[bianma+"areablod"]
    #         up_loaddir = up_dic.replace("/files", "/mnt/uploaded_files")
    #         upload(bianma_pic, up_loaddir)
    #         print(f"{bianma_pic}水田图片上传成功_{bianma_pic}")
    #     if tudileixin == "旱地":
    #         bianma_pic = yumi_pic_list[random.randint(0, len(yumi_pic_list))-1]
    #         up_dic = dict_adress[bianma + "areablod"]
    #         up_loaddir = up_dic.replace("/files", "/mnt/uploaded_files")
    #         upload(bianma_pic, up_loaddir)
    #         print(f"{bianma_pic}旱地图片上传成功_{bianma_pic}")
    n =0
    for one_row in range(2, rows+1):
        bianma = ws[f"A{one_row}"].value
        pinzhong = ws[f"D{one_row}"].value
        if pinzhong in pinzhong_dict:
            bianma_pic = pinzhong_dict[pinzhong]
            try:
                up_dic = dict_adress[bianma + "variety"]
                up_loaddir = up_dic.replace("/files", "/mnt/uploaded_files")
                upload(bianma_pic, up_loaddir)
                print(f"{bianma_pic}品种图片上传成功_{bianma_pic}")
            except:
                print("无图")
                n+=1
        else:
            bianma_pic = r"C:\Users\65680\Desktop\pinzhogn.jpg"
            try:
                up_dic = dict_adress[bianma + "variety"]
                up_loaddir = up_dic.replace("/files", "/mnt/uploaded_files")
                upload(bianma_pic, up_loaddir)
                print(f"{bianma_pic}品种图片上传成功_{bianma_pic}")
            except:
                n+=1
print(n)

