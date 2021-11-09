#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Date:2020-12-18
# Author:Runker54
# -----------------------
import sys
from PyQt5.QtWidgets import QMainWindow, QApplication, QFileDialog
# from PyQt5.QtGui import *
import openpyxl
import os
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QTimer, QDateTime
import random


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(800, 600)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        # 按钮
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)  # 编码转换按钮
        self.pushButton.setGeometry(QtCore.QRect(20, 320, 100, 50))
        self.pushButton.setObjectName("pushButton")
        # self.pushButton_2 = QtWidgets.QPushButton(self.centralwidget)  # 等级评价按钮
        # self.pushButton_2.setGeometry(QtCore.QRect(20, 390, 100, 50))
        # self.pushButton_2.setObjectName("pushButton_2")
        # self.pushButton_3 = QtWidgets.QPushButton(self.centralwidget)  # 合并单个book多个sheet按钮
        # self.pushButton_3.setGeometry(QtCore.QRect(140, 320, 100, 50))
        # self.pushButton_3.setObjectName("pushButton_3")
        # self.pushButton_4 = QtWidgets.QPushButton(self.centralwidget)  # 合并单个book多个sheet按钮
        # self.pushButton_4.setGeometry(QtCore.QRect(140, 390, 100, 50))
        # self.pushButton_4.setObjectName("pushButton_4")
        # 标签
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(20, 40, 72, 15))
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(430, 40, 72, 15))
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.centralwidget)   # 输出详细信息
        self.label_3.setGeometry(QtCore.QRect(430, 300, 101, 16))  # 输出详细信息
        self.label_3.setObjectName("label_3")  # 输出详细信息
        self.label_4 = QtWidgets.QLabel(self.centralwidget)   # at信息
        self.label_4.setGeometry(QtCore.QRect(20, 560, 250, 20))  # at信息  x, y, length, height
        self.label_4.setObjectName("label_4")  # at信息
        # 文本框
        self.textEdit = QtWidgets.QTextEdit(self.centralwidget)
        self.textEdit.setGeometry(QtCore.QRect(20, 70, 341, 211))
        self.textEdit.setObjectName("textEdit")
        self.textEdit_2 = QtWidgets.QTextEdit(self.centralwidget)
        self.textEdit_2.setGeometry(QtCore.QRect(430, 70, 341, 211))
        self.textEdit_2.setObjectName("textEdit_2")
        self.textEdit_3 = QtWidgets.QTextEdit(self.centralwidget)
        self.textEdit_3.setGeometry(QtCore.QRect(430, 330, 341, 211))
        self.textEdit_3.setObjectName("textEdit_3")

        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 800, 26))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.retranslateUi(MainWindow)
        # 触发信号
        self.pushButton.clicked.connect(MainWindow.id_change)
        # self.pushButton_2.clicked.connect(MainWindow.level_calc)
        # self.pushButton_3.clicked.connect(MainWindow.merge_sheet)
        # self.pushButton_4.clicked.connect(MainWindow.chose_directory)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))

        self.label.setText(_translate("MainWindow", "源文件："))
        self.label_2.setText(_translate("MainWindow", "目标文件："))
        self.label_3.setText(_translate("MainWindow", "提示信息："))
        self.label_4.setText(_translate("MainWindow", "Runker54"))

        self.pushButton.setText(_translate("MainWindow", "数据评价"))
        # self.pushButton_2.setText(_translate("MainWindow", "等级评价"))
        # self.pushButton_3.setText(_translate("MainWindow", "单个表格合并"))
        # self.pushButton_4.setText(_translate("MainWindow", "多个表格合并"))


class DetailUI(Ui_MainWindow, QMainWindow):
    def __init__(self):
        super(DetailUI, self).__init__()
        self.setupUi(self)
        self.setWindowTitle('农产品评价')
        self.textEdit_3.setText("")

    def id_change(self):
        # 限值字典
        cache_dict = {"水稻": {"Cd": 0.2, "Hg": 0.02, "As": 0.5, "Pb": 0.2, "Cr": 1},
                      "玉米": {"Cd": 0.1, "Hg": 0.02, "As": 0.5, "Pb": 0.2, "Cr": 1},
                      "马铃薯": {"Cd": 0.1, "Hg": 0.01, "As": 0.5, "Pb": 0.2, "Cr": 0.5},
                      "薏仁米": {"Cd": 0.1, "Hg": 0.02, "As": 0.5, "Pb": 0.2, "Cr": 1},
                      "小麦": {"Cd": 0.1, "Hg": 0.02, "As": 0.5, "Pb": 0.2, "Cr": 1},
                      "豆类": {"Cd": 0.1, "Hg": 0.01, "As": 0.5, "Pb": 0.2, "Cr": 1},
                      "高粱": {"Cd": 0.1, "Hg": 0.02, "As": 0.5, "Pb": 0.2, "Cr": 1}}
        # 相对偏差字典
        xdpc_dict = {"Cd": [0.1, 35, 35, 0.2, 30, 25],
                     "Hg": [0.1, 35, 35, 0.2, 30, 25],
                     "As": [0.1, 35, 35, 0.2, 30, 25],
                     "Pb": [0.1, 35, 35, 1.0, 30, 25],
                     "Cr": [0.1, 35, 35, 1.0, 30, 25]}
        try:
            fname, _ = QFileDialog.getOpenFileName(self, 'open file', '/', "Excel files (*.xls *.xlsx)")
            fname1, _ = QFileDialog.getOpenFileName(self, 'open file', '/', "Excel files (*.xls *.xlsx)")
            self.textEdit.setText(fname)
            xls_path = fname  # 得到文件路径
            xls_path1 = fname1
            temp_path = os.path.splitext(xls_path)
            new_path = temp_path[0] + "_等级评价" + temp_path[1]

            w_b = openpyxl.load_workbook(xls_path)
            w_b_zk = openpyxl.load_workbook(xls_path1)
            w_s_zk = w_b_zk.active
            w_s = w_b.active
            # 二次编码对应原始编码数据收集 变量名 erci_bianma  是否质控点位收集  变量名 zk_dict
            erci_bianma = {}
            zk_dict = {}
            b_rows = w_s_zk.max_row
            b_col = w_s_zk.max_column
            for zk_row in range(2, b_rows + 1):
                erci_bianma[f"{w_s_zk.cell(zk_row, 2).value}"] = w_s_zk.cell(zk_row, 1).value
                if w_s_zk.cell(zk_row, 3).value is not None:
                    zk_dict[f"{w_s_zk.cell(zk_row, 2).value}"] = w_s_zk.cell(zk_row, 3).value
            title_list = ["Cd_level", "Hg_level", "As_level", "Pb_level", "Cr_level", "原始编码"]
            r = 13
            # 写入抬头
            for _i in title_list:
                w_s.cell(1, r).value = _i
                r += 1
            rows = w_s.max_row
            cols = w_s.max_column
            level_data_dict = {}
            zk_list = []
            for one_row in range(2, rows + 1):
                # 写入原始编码
                w_s.cell(one_row, 18).value = str(erci_bianma[w_s.cell(one_row, 4).value])
                # 取值
                cd_v = w_s[f"F{one_row}"].value if type(w_s[f"F{one_row}"].value) is not str else 0  # 镉
                hg_v = w_s[f"G{one_row}"].value if type(w_s[f"G{one_row}"].value) is not str else 0  # 汞
                as_v = w_s[f"H{one_row}"].value if type(w_s[f"H{one_row}"].value) is not str else 0  # 砷
                pb_v = w_s[f"I{one_row}"].value if type(w_s[f"I{one_row}"].value) is not str else 0  # 铅
                cr_v = w_s[f"J{one_row}"].value if type(w_s[f"J{one_row}"].value) is not str else 0  # 镉

                s_v = {"Cd": cd_v, "Hg": hg_v, "As": as_v, "Pb": pb_v, "Cr": cr_v}
                yplb = w_s[f"E{one_row}"].value  # 样品类型
                cd_mv = cache_dict[yplb]["Cd"]  # 镉值
                hg_mv = cache_dict[yplb]["Hg"]  # 汞值
                as_mv = cache_dict[yplb]["As"]  # 铅值
                pb_mv = cache_dict[yplb]["Pb"]  # 砷值
                cr_mv = cache_dict[yplb]["Cr"]  # 铬值
                s_mv = {"Cd": cd_mv, "Hg": hg_mv, "As": as_mv, "Pb": pb_mv, "Cr": cr_mv}
                level = 0
                #  判断
                rows_ = 13
                for index_, one_cell in enumerate(["Cd", "Hg", "As", "Pb", "Cr"]):
                    if s_v[one_cell] <= s_mv[one_cell]:
                        level = 1
                    elif s_v[one_cell] <= 2 * s_mv[one_cell]:
                        level = 2
                    else:
                        level = 3
                    w_s.cell(one_row, rows_).value = level
                    rows_ += 1
                # 搜集等级数据、包名、值等  变量名 level_data_dict   质控数据变量名  zk_list

                for one_data in zk_dict:
                    if w_s.cell(one_row, 4).value == one_data:
                        level_data_dict[one_data] = [w_s.cell(one_row, 13).value,
                                                     w_s.cell(one_row, 14).value,
                                                     w_s.cell(one_row, 15).value,
                                                     w_s.cell(one_row, 16).value,
                                                     w_s.cell(one_row, 17).value,
                                                     w_s.cell(one_row, 6).value,
                                                     w_s.cell(one_row, 7).value,
                                                     w_s.cell(one_row, 8).value,
                                                     w_s.cell(one_row, 9).value,
                                                     w_s.cell(one_row, 10).value,
                                                     w_s.cell(one_row, 1).value,
                                                     w_s.cell(one_row, 18).value,
                                                     w_s.cell(one_row, 4).value,
                                                     zk_dict[w_s.cell(one_row, 4).value]]
                        zk_list.append(level_data_dict[one_data])
            bao_list = []
            for one_element in level_data_dict:
                bao_list.append(level_data_dict[one_element][10])
            bao_xunhuan_list = list(set(bao_list))
            px_s = w_b.create_sheet("平行样质控")
            bz_s = w_b.create_sheet("标准物质质控")
            # 写入该表抬头
            px_title = ["项目", "cd_level", "hg_level", "as_level", "pb_level", "cr_level", "cd_level", "hg_level",
                        "as_level", "pb_level", "cr_level", "包名", "原始编码", "二次编码", "样品类型"]
            for _number, one_i in enumerate(px_title):
                px_s.cell(1, _number + 1).value = one_i
                bz_s.cell(1, _number + 1).value = one_i
            bz_r = 2  # 标准物质行控制
            px_r = 2  # 平行样行控制
            for one_bao in bao_xunhuan_list:  # 按包名循环值
                rr = 0
                for one_list in zk_list:  # 在数据列表中寻找
                    if one_list[10] == one_bao:  # 数据列表中找到指定包名数据
                        if one_list[13].strip() != "标准物质":  # 标准物质数据写入及判断
                            rr += 1
                            for px_number, px_row in enumerate(one_list):
                                px_s.cell(px_r, px_number + 2).value = px_row
                            if rr == 2:
                                px_r += 3
                                px_s.cell(px_r - 2, 1).value = "一致性及相对偏差"
                                # 相对偏差对比  收集两列的数值  upv_list  downv_list
                                upv_list = [px_s.cell(px_r - 4, _upv).value for _upv in range(7, 12)]
                                downv_list = [px_s.cell(px_r - 3, _downv).value for _downv in range(7, 12)]
                                for upv_number, onev_up in enumerate(upv_list):
                                    try:
                                        pc_v = abs(onev_up - downv_list[upv_number]) / (
                                                    onev_up + downv_list[upv_number]) * 100
                                        px_s.cell(px_r - 2, upv_number + 7).value = pc_v
                                    except:
                                        px_s.cell(px_r - 2, upv_number + 7).value = "\\"
                                # 一致性对比  收集两列的等级值  up_list  down_list
                                up_list = [px_s.cell(px_r - 4, _up).value for _up in range(2, 7)]
                                down_list = [px_s.cell(px_r - 3, _down).value for _down in range(2, 7)]
                                for up_number, one_up in enumerate(up_list):
                                    if down_list[up_number] != one_up:
                                        px_s.cell(px_r - 2, up_number + 2).value = "不一致"
                                        pc_mv = px_s.cell(px_r - 2, up_number + 7).value

                                        yuansu_dict = {2: "Cd", 3: "Hg", 4: "As", 5: "Pb", 6: "Cr"}
                                        yuansu = yuansu_dict[up_number + 2]  # 判定为什么元素
                                        yuansuzhi = max([px_s.cell(px_r - 3, up_number + 7).value,
                                                         px_s.cell(px_r - 4, up_number + 7).value])  # 两个元素的最大值
                                        # 与第一个对比
                                        pc_xianzhi = 0
                                        if yuansuzhi < xdpc_dict[yuansu][0]:
                                            pc_xianzhi = xdpc_dict[yuansu][1]
                                        elif yuansuzhi < xdpc_dict[yuansu][3]:
                                            pc_xianzhi = xdpc_dict[yuansu][4]
                                        else:
                                            pc_xianzhi = xdpc_dict[yuansu][5]
                                        if pc_mv < pc_xianzhi:
                                            px_s.cell(px_r - 1, up_number + 7).value = "合格"
                                        else:
                                            px_s.cell(px_r - 1, up_number + 7).value = "不合格"
                                    else:
                                        px_s.cell(px_r - 2, up_number + 2).value = "一致"
                                        px_s.cell(px_r - 1, up_number + 7).value = "合格"
                                px_s.cell(px_r - 1, 1).value = "是否合格"
                            else:
                                px_r += 1
                        else:
                            for bz_number, bz_row in enumerate(one_list):
                                bz_s.cell(bz_r, bz_number + 2).value = bz_row
                            bz_r += 2
            w_b.save(new_path)
            # new_work_book.save(new_path)
            self.textEdit.setText("所选择文件:%s处理完成！" % fname)
            self.textEdit_2.setText("输出文件路径:%s" % new_path)
        except:
            self.textEdit.setText("所选择文件类型有误或表格格式不正确！")




if __name__ == "__main__":
    app = QApplication(sys.argv)
    ex = DetailUI()
    ex.show()
    sys.exit(app.exec_())
