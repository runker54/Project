#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Date:2021-11-17
# Author:Runker54
# ----------------------
import openpyxl
import os

# 农产品限值字典
select_agri = {"水稻": {"Cd": 0.2, "Hg": 0.02, "As": 0.5, "Pb": 0.2, "Cr": 1},
               "玉米": {"Cd": 0.1, "Hg": 0.02, "As": 0.5, "Pb": 0.2, "Cr": 1},
               "马铃薯": {"Cd": 0.1, "Hg": 0.01, "As": 0.5, "Pb": 0.2, "Cr": 0.5},
               "薏仁米": {"Cd": 0.1, "Hg": 0.02, "As": 0.5, "Pb": 0.2, "Cr": 1},
               "小麦": {"Cd": 0.1, "Hg": 0.02, "As": 0.5, "Pb": 0.2, "Cr": 1},
               "豆类": {"Cd": 0.1, "Hg": 0.01, "As": 0.5, "Pb": 0.2, "Cr": 1},
               "高粱": {"Cd": 0.1, "Hg": 0.02, "As": 0.5, "Pb": 0.2, "Cr": 1},
               "辣椒": {"Cd": 0.05, "Hg": 0.01, "As": 0.5, "Pb": 0.2, "Cr": 1},
               "红薯": {"Cd": 0.1, "Hg": 0.01, "As": 0.5, "Pb": 0.2, "Cr": 0.5},
               "茄子": {"Cd": 0.05, "Hg": 0.01, "As": 0.5, "Pb": 0.1, "Cr": 0.5}}
# 土壤限制
select_soil = {'水田': {5: {'Cd': [0.3, 1.5], 'Hg': [0.5, 2.0], 'As': [30, 200], 'Pb': [80, 400], 'Cr': [250, 800]},
                      6: {'Cd': [0.4, 2.0], 'Hg': [0.5, 2.5], 'As': [30, 150], 'Pb': [100, 500], 'Cr': [250, 850]},
                      7: {'Cd': [0.6, 3.0], 'Hg': [0.6, 4.0], 'As': [25, 120], 'Pb': [140, 700], 'Cr': [300, 1000]},
                      8: {'Cd': [0.8, 4.0], 'Hg': [1.0, 6.0], 'As': [20, 100], 'Pb': [240, 1000], 'Cr': [350, 1300]}},
               '其它': {5: {'Cd': [0.3, 1.5], 'Hg': [1.3, 2.0], 'As': [40, 200], 'Pb': [70, 400], 'Cr': [150, 800]},
                      6: {'Cd': [0.3, 2.0], 'Hg': [1.8, 2.5], 'As': [40, 150], 'Pb': [90, 500], 'Cr': [150, 850]},
                      7: {'Cd': [0.3, 3.0], 'Hg': [2.4, 4.0], 'As': [30, 120], 'Pb': [120, 700], 'Cr': [200, 1000]},
                      8: {'Cd': [0.6, 4.0], 'Hg': [3.4, 6.0], 'As': [25, 100], 'Pb': [170, 1000], 'Cr': [250, 1300]}}}
# 相对偏差字典
xdpc_dict = {"Cd": [0.1, 35, 35, 0.2, 30, 25],
             "Hg": [0.1, 35, 35, 0.2, 30, 25],
             "As": [0.1, 35, 35, 0.2, 30, 25],
             "Pb": [0.1, 35, 35, 1.0, 30, 25],
             "Cr": [0.1, 35, 35, 1.0, 30, 25],
             }

fdir_path = r"C:\Users\65680\Desktop\test"
data_shuju_path = r"C:\Users\65680\Desktop\test\1纳雍县数据.xlsx"  # 数据表路径
data_bianma_path = r"C:\Users\65680\Desktop\test\2纳雍县质控点位表.xlsx"  # 质控点位表路径
data_biaozhun_path = r"C:\Users\65680\Desktop\test\3标准物质.xlsx"  # 标准物质路径
w_b = openpyxl.load_workbook(data_shuju_path)
w_b_zk = openpyxl.load_workbook(data_bianma_path)
w_s_zk = w_b_zk.active
w_s = w_b.active


# 判断各元素等级函数
def c_level(lx, value, yuansu):
    """传入该点的作物类型，对应元素值，对应的元素，返回该点的等级"""
    try:
        value = float(value)
    except:
        pass
    value1 = float(value) if type(value) is not str else 0
    if yuansu not in ["Cd", "Hg", "As", "Pb", "Cr"]:
        level = None
    else:
        if lx in select_agri:
            xianzhi = select_agri[lx][yuansu]
            if value1 <= xianzhi:
                level = 1
            elif value1 <= 2 * xianzhi:
                level = 2
            else:
                level = 3
        else:
            level = None
    return level


# 收集标准物质信息
def data_bz(path):
    """"返回{"农产品":[],"土壤":[]}字典数据"""
    bz_dict_n = {}
    bz_dict_t = {}
    w_b_bz = openpyxl.load_workbook(data_biaozhun_path)
    w_s_bz_n = w_b_bz["农产品"]
    w_s_bz_t = w_b_bz["土壤"]
    rows_n = w_s_bz_n.max_row
    for one_n_row in range(2, rows_n + 1):
        yuansu_n = list(list(w_s_bz_n.rows)[one_n_row - 1])
        bz_dict_n[yuansu_n[0].value] = [bz_n.value for bz_n in yuansu_n[1:]]
    rows_t = w_s_bz_t.max_row
    for one_t_row in range(2, rows_t + 1):
        yuansu_t = list(list(w_s_bz_t.rows)[one_t_row - 1])
        bz_dict_t[yuansu_t[0].value] = [bz_t.value for bz_t in yuansu_t[1:]]
    bz_dict = {"农产品": bz_dict_n, "土壤": bz_dict_t}
    return bz_dict


# 收集二次编码对应的原始编码
def data_ercibianma(path):
    """收集二次编码对应的原始编码字典erci_dict = {样品编号：原始编码}"""
    erci_dict = {}
    work_book = openpyxl.load_workbook(path)
    ws = work_book.active
    rows = ws.max_row
    for _row in range(2, rows + 1):
        erci_dict[ws.cell(_row, 2).value] = ws.cell(_row, 1).value
    return erci_dict


# 收集那些编码属于平行样品和标准物质
def data_pbbianma(path):
    """返回标准物质和室内时间平行的数据字典pb_dict{"标准物质":[],"平行样品"：[],}"""
    bz, px = [], []
    work_book = openpyxl.load_workbook(path)
    ws = work_book.active
    rows = ws.max_row
    bz_number_list = []
    px_dict = {}
    bz_dict = {}
    for _row in range(2, rows + 1):
        if "标准" in str(ws.cell(_row, 3).value).strip():
            b_number = check_data[ws.cell(_row, 2).value][0]
            bz_number_list.append(b_number)
            bz.append([ws.cell(_row, 2).value, b_number])
        elif "平行" in str(ws.cell(_row, 3).value).strip():
            b_number = check_data[ws.cell(_row, 2).value][0]
            bz_number_list.append(b_number)
            px.append([ws.cell(_row, 2).value, b_number])
        else:
            pass
    for _one_number in bz_number_list:
        _cahcepx = []
        _cahcebz = []
        for _px in px:
            if _px[1] == _one_number:
                _cahcepx.append(_px[0])
                px_dict[_one_number] = _cahcepx
        for _bz in bz:
            if _bz[1] == _one_number:
                _cahcebz.append(_bz[0])
                bz_dict[_one_number] = _cahcebz
    pb_dict = {"标准物质": bz_dict, "平行样品": px_dict}
    return pb_dict


# 收集检测数据及包信息
def data_check(path):
    """收集检测数据信息，以字典形式存放数据，check_dict = {"样品编号"：[所有列的数据]}"""
    check_dict = {}
    work_book = openpyxl.load_workbook(path)
    ws = work_book.active
    rows = ws.max_row
    cols = ws.max_column
    for _row in range(2, rows + 1):
        _list = []
        for _col in range(1, cols):
            _list.append(ws.cell(_row, _col).value)
        check_dict[ws.cell(_row, 4).value] = _list
    return check_dict


# 标准物质信息
bz_data = data_bz(data_biaozhun_path)
# 二次编码对应原始编码信息
erci_data = data_ercibianma(data_bianma_path)
# 检测数据表信息
check_data = data_check(data_shuju_path)
# 平行样平和标准物质信息
pxbz_data = data_pbbianma(data_bianma_path)


# 写入判断后的结果数据
def write_message(path):
    """"依据所得数据开始往表格内写入"""
    work_book = openpyxl.load_workbook(path)
    ws = work_book.active
    rows = ws.max_row
    cols = ws.max_column
    title = ["原始编码", "Cd_level", "Hg_level", "As_level", "Pb_level", "Cr_level", "Se_level", "Ge_level", "Zh_level"]
    # 写入抬头
    for _add, _title in enumerate(title):
        ws.cell(1, cols + 1 + _add).value = _title
    # 写入抬头对应的数据
    for _row in range(2, rows + 1):
        # 写入原始编码
        try:
            ws.cell(_row, cols + 1).value = str(erci_data[ws.cell(_row, 4).value])
        except:
            pass
        # 写入各值对应的等级
        one_yuansu_list = []
        for _id, _yuansu in enumerate(["Cd", "Hg", "As", "Pb", "Cr", "Se", "Ge"]):
            ws.cell(_row, cols + 2 + _id).value = c_level(ws.cell(_row, 5).value, ws.cell(_row, 6 + _id).value, _yuansu)
            one_yuansu_list.append(c_level(ws.cell(_row, 5).value, ws.cell(_row, 6 + _id).value, _yuansu))
        one_yuansu_list = list(set(one_yuansu_list))
        one_yuansu_list.remove(None)
        try:
            ws.cell(_row, cols + 9).value = max(one_yuansu_list)
        except:
            pass
    # 写平行样
    ws_px = work_book.create_sheet("平行样质控")
    title_px_bz = ["项目", "原始编码", "样包号", "内部编号", "样品编号", "样品名称", "Cd_level", "Hg_level", "As_level", "Pb_level",
                   "Cr_level", "Cd", "Hg", "As", "Pb", "Cr", "Se", "Ge"]
    for _id, _title in enumerate(title_px_bz):
        ws_px.cell(1, _id+1).value = _title
    px_message = pxbz_data["平行样品"]
    for one_bao in px_message:
        for one_index in pxbz_data["平行样品"][one_bao]:
            print(check_data)
            one_mesage = [erci_data[one_index],
                          check_data[one_index][0],
                          check_data[one_index][1],
                          check_data[one_index][3],
                          check_data[one_index][4], ]
            print(one_mesage)
    # 写标准物质
    ws_bz = work_book.create_sheet("标准物质质控")
    for _id, _title in enumerate(title_px_bz):
        ws_bz.cell(1, _id+1).value = _title
    return work_book


write_message(data_shuju_path)
# write_message(data_shuju_path).save(r"C:\Users\65680\Desktop\test.xlsx")

# px_s = w_b.create_sheet("平行样质控")
# bz_s = w_b.create_sheet("标准物质质控")
# # 写入该表抬头
# px_title = ["项目", "cd_level", "hg_level", "as_level", "pb_level", "cr_level", "se_level", "ge_level", "cd_value",
#             "hg_value", "as_value", "pb_value", "cr_value", "se_value", "ge_value", "包名", "原始编码", "二次编码", "样品类型"]
# for _number, one_i in enumerate(px_title):
#     px_s.cell(1, _number + 1).value = one_i
#     bz_s.cell(1, _number + 1).value = one_i
# bz_r = 2  # 标准物质行控制
# px_r = 2  # 平行样行控制
# for one_bao in bao_xunhuan_list:  # 按包名循环值
#     rr = 0
#     for one_list in zk_list:  # 在数据列表中寻找
#         if one_list[14] == one_bao:  # 数据列表中找到指定包名数据
#             if one_list[17].strip() != "标准物质":  # 标准物质数据写入及判断
#                 rr += 1
#                 for px_number, px_row in enumerate(one_list):
#                     px_s.cell(px_r, px_number + 2).value = px_row
#                 if rr == 2:
#                     px_r += 3
#                     px_s.cell(px_r - 2, 1).value = "一致性及相对偏差"
#                     # 相对偏差对比  收集两列的数值  upv_list  downv_list
#                     upv_list = [px_s.cell(px_r - 4, _upv).value for _upv in range(9, 16)]
#                     downv_list = [px_s.cell(px_r - 3, _downv).value for _downv in range(9, 16)]
#                     for upv_number, onev_up in enumerate(upv_list):
#                         try:
#                             pc_v = abs(onev_up - downv_list[upv_number]) / (onev_up + downv_list[upv_number]) * 100
#                             px_s.cell(px_r - 2, upv_number + 9).value = pc_v
#                         except:
#                             px_s.cell(px_r - 2, upv_number + 9).value = "\\"
#                     # 一致性对比  收集两列的等级值  up_list  down_list
#                     up_list = [px_s.cell(px_r - 4, _up).value for _up in range(2, 9)]
#                     down_list = [px_s.cell(px_r - 3, _down).value for _down in range(2, 9)]
#                     for up_number, one_up in enumerate(up_list):
#                         if down_list[up_number] != one_up:
#                             px_s.cell(px_r - 2, up_number + 2).value = "不一致"
#                             pc_mv = px_s.cell(px_r - 2, up_number + 9).value
#                             yuansu_dict = {2: "Cd", 3: "Hg", 4: "As", 5: "Pb", 6: "Cr", 7: "Se", 8: "Ge"}
#                             yuansu = yuansu_dict[up_number + 2]  # 判定为什么元素
#
#                             yuansuzhi = max([px_s.cell(px_r - 3, up_number + 9).value,
#                                              px_s.cell(px_r - 4, up_number + 9).value])  # 两个元素的最大值
#                             # 与第一个对比
#                             pc_xianzhi = 0
#                             if yuansuzhi < xdpc_dict[yuansu][0]:
#                                 pc_xianzhi = xdpc_dict[yuansu][1]
#                             elif yuansuzhi < xdpc_dict[yuansu][3]:
#                                 pc_xianzhi = xdpc_dict[yuansu][4]
#                             else:
#                                 pc_xianzhi = xdpc_dict[yuansu][5]
#
#                             if pc_mv < pc_xianzhi:
#                                 px_s.cell(px_r - 1, up_number + 9).value = "合格"
#                             else:
#                                 px_s.cell(px_r - 1, up_number + 9).value = "不合格"
#                         else:
#                             px_s.cell(px_r - 2, up_number + 2).value = "一致"
#                             px_s.cell(px_r - 1, up_number + 9).value = "合格"
#                     px_s.cell(px_r - 1, 1).value = "是否合格"
#                 else:
#                     px_r += 1
#             else:
#                 # 标准物质编号
#                 bz_s.cell(bz_r + 1, 1).value = "标准物质含量"
#                 bz_s.cell(bz_r, 1).value = "原始值"
#                 bz_bh = one_list[15]
#                 bz_data = bz_message["农产品"][bz_bh][:7]
#                 # 写入标准物质检测值
#                 for bz_number, bz_row in enumerate(one_list):
#                     bz_s.cell(bz_r, bz_number + 2).value = bz_row
#                 bz_s.cell(bz_r + 2, 1).value = "相对偏差值"
#                 bz_s.cell(bz_r + 3, 1).value = "是否合格"
#                 # 写入标准物质参考值
#                 for one_bz, bz_nn in enumerate(bz_data):
#
#                     # 对应元素值
#                     yuansu_vv = bz_s.cell(bz_r, one_bz + 9).value if type(
#                         bz_s.cell(bz_r, one_bz + 9).value) is not str else 0
#                     # 标准物质值
#                     bz_s.cell(bz_r + 1, one_bz + 9).value = bz_nn
#                     # 判断为何种元素
#                     yuansu_bvdict = {2: "Cd", 3: "Hg", 4: "As", 5: "Pb", 6: "Cr", 7: "Se", 8: "Ge"}
#                     yuansu_bv = yuansu_bvdict[one_bz + 2]
#                     if yuansu_bv in ["Se", "Ge"]:
#                         bz_s.cell(bz_r + 3, one_bz + 9).value = "-"
#                     else:
#                         # 元素值的相对偏差参考值获取
#                         pcbv_xianzhi = 0
#                         if yuansu_vv < xdpc_dict[yuansu_bv][0]:
#                             pcbv_xianzhi = xdpc_dict[yuansu_bv][1]
#                         elif yuansu_vv < xdpc_dict[yuansu_bv][3]:
#                             pcbv_xianzhi = xdpc_dict[yuansu_bv][4]
#                         else:
#                             pcbv_xianzhi = xdpc_dict[yuansu_bv][5]
#                         # 判断使用何种方法做计算
#                         # 第一种含 ± 号值
#                         if "±" in str(bz_nn):
#                             left_v = float(str(bz_nn)[:str(bz_nn).find("±")])
#                             right_v = float(str(bz_nn)[str(bz_nn).find("±") + 1:])
#                             up_bv, down_bv = left_v - right_v, left_v + right_v
#                             # 元素值与前值的相对偏差
#                             bz_xdpc = abs(yuansu_vv - left_v) / (yuansu_vv + left_v) * 100
#                             bz_s.cell(bz_r + 2, one_bz + 9).value = bz_xdpc
#                             # 判断方法，判定是否在范围值内,或判定与前值的相对偏差是否合理
#                             if ((yuansu_vv <= up_bv) and (yuansu_vv >= down_bv)) or (bz_xdpc < pcbv_xianzhi):
#                                 bz_s.cell(bz_r + 3, one_bz + 9).value = "合格"
#                             else:
#                                 bz_s.cell(bz_r + 3, one_bz + 9).value = "不合格"
#
#                         # 第二中含（）值
#                         if "-" in str(bz_nn):
#                             # 参考值
#                             bz_bv = float(str(bz_nn).strip("-"))
#                             # 相对偏差值
#                             bz_xdpcer = abs(yuansu_vv - bz_bv) / (yuansu_vv + bz_bv) * 100
#                             bz_s.cell(bz_r + 2, one_bz + 9).value = bz_xdpcer
#                             if bz_xdpcer < pcbv_xianzhi:
#                                 bz_s.cell(bz_r + 3, one_bz + 9).value = "合格"
#                             else:
#                                 bz_s.cell(bz_r + 3, one_bz + 9).value = "不合格"
#                 bz_r += 4
#
# w_b.save(os.path.join(fdir_path, "等级判定结果.xlsx"))
