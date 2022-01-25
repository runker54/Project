#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Date:2021-10-28
# Author:Runker54
# ----------------------

import openpyxl
import os

# 限值字典
cache_dict = {"水稻": {"Cd": 0.2, "Hg": 0.02, "As": 0.5, "Pb": 0.2, "Cr": 1},
              "玉米": {"Cd": 0.1, "Hg": 0.02, "As": 0.5, "Pb": 0.2, "Cr": 1},
              "马铃薯": {"Cd": 0.1, "Hg": 0.01, "As": 0.5, "Pb": 0.2, "Cr": 0.5},
              "薏仁米": {"Cd": 0.1, "Hg": 0.02, "As": 0.5, "Pb": 0.2, "Cr": 1},
              "小麦": {"Cd": 0.1, "Hg": 0.02, "As": 0.5, "Pb": 0.2, "Cr": 1},
              "豆类": {"Cd": 0.1, "Hg": 0.01, "As": 0.5, "Pb": 0.2, "Cr": 1},
              "高粱": {"Cd": 0.1, "Hg": 0.02, "As": 0.5, "Pb": 0.2, "Cr": 1},
              "辣椒": {"Cd": 0.05, "Hg": 0.01, "As": 0.5, "Pb": 0.2, "Cr": 1},
              "红薯": {"Cd": 0.1, "Hg": 0.01, "As": 0.5, "Pb": 0.2, "Cr": 0.5},
              "茄子": {"Cd": 0.05, "Hg": 0.01, "As": 0.5, "Pb": 0.1, "Cr": 0.5}}
# 相对偏差字典
xdpc_dict = {"Cd": [0.1, 35, 35, 0.2, 30, 25],
             "Hg": [0.1, 35, 35, 0.2, 30, 25],
             "As": [0.1, 35, 35, 0.2, 30, 25],
             "Pb": [0.1, 35, 35, 1.0, 30, 25],
             "Cr": [0.1, 35, 35, 1.0, 30, 25],
             }

fdir_path = r"E:\GisWorkSpace\提交资料20211111\从江县\从江县检测数据_农产品_1126"
data_shuju_path = r"E:\GisWorkSpace\提交资料20211111\从江县\从江县检测数据_农产品_1126\1从江县数据.xlsx"  # 数据表路径
data_bianma_path = r"E:\GisWorkSpace\提交资料20211111\从江县\从江县检测数据_农产品_1126\2质控点位表.xlsx"  # 质控点位表路径
data_biaozhun_path = r"E:\GisWorkSpace\提交资料20211111\从江县\从江县检测数据_农产品_1126\3标准物质.xlsx"  # 标准物质路径
w_b = openpyxl.load_workbook(data_shuju_path)
w_b_zk = openpyxl.load_workbook(data_bianma_path)
w_s_zk = w_b_zk.active
w_s = w_b.active


# 收集标准物质信息
def data_bz(path):
    """"返回{"农产品":[],"土壤":[]}字典数据"""
    bz_dict_n = {}
    bz_dict_t = {}
    w_b_bz = openpyxl.load_workbook(data_biaozhun_path)
    w_s_bz_n = w_b_bz["农产品"]
    w_s_bz_t = w_b_bz["土壤"]
    rows_n = w_s_bz_n.max_row
    for one_n_row in range(2, rows_n + 1):
        yuansu_n = list(list(w_s_bz_n.rows)[one_n_row - 1])
        bz_dict_n[yuansu_n[0].value] = [bz_n.value for bz_n in yuansu_n[1:]]
    rows_t = w_s_bz_t.max_row
    for one_t_row in range(2, rows_t + 1):
        yuansu_t = list(list(w_s_bz_t.rows)[one_t_row - 1])
        bz_dict_t[yuansu_t[0].value] = [bz_t.value for bz_t in yuansu_t[1:]]
    bz_dict = {"农产品": bz_dict_n, "土壤": bz_dict_t}
    return bz_dict


# 标准物质信息
bz_message = data_bz(data_biaozhun_path)

# 二次编码对应原始编码数据收集 变量名 erci_bianma  是否质控点位收集  变量名 zk_dict
erci_bianma = {}
zk_dict = {}
b_rows = w_s_zk.max_row
b_col = w_s_zk.max_column
for zk_row in range(2, b_rows + 1):
    erci_bianma[f"{w_s_zk.cell(zk_row, 2).value}"] = w_s_zk.cell(zk_row, 1).value
    if w_s_zk.cell(zk_row, 3).value is not None:
        zk_dict[f"{w_s_zk.cell(zk_row, 2).value}"] = w_s_zk.cell(zk_row, 3).value
title_list = ["Cd_level", "Hg_level", "As_level", "Pb_level", "Cr_level", "Se_level", "Ge_level", "原始编码"]
r = 13
# 写入抬头
for _i in title_list:
    w_s.cell(1, r).value = _i
    r += 1
rows = w_s.max_row
cols = w_s.max_column
level_data_dict = {}
zk_list = []
for one_row in range(2, rows + 1):
    # 写入原始编码
    w_s.cell(one_row, 20).value = str(erci_bianma[w_s.cell(one_row, 4).value])
    # 取值
    cd_v = w_s[f"F{one_row}"].value if type(w_s[f"F{one_row}"].value) is not str else 0  # 镉
    hg_v = w_s[f"G{one_row}"].value if type(w_s[f"G{one_row}"].value) is not str else 0  # 汞
    as_v = w_s[f"H{one_row}"].value if type(w_s[f"H{one_row}"].value) is not str else 0  # 砷
    pb_v = w_s[f"I{one_row}"].value if type(w_s[f"I{one_row}"].value) is not str else 0  # 铅
    cr_v = w_s[f"J{one_row}"].value if type(w_s[f"J{one_row}"].value) is not str else 0  # 镉
    se_v = w_s[f"J{one_row}"].value if type(w_s[f"J{one_row}"].value) is not str else 0  # 硒
    ge_v = w_s[f"J{one_row}"].value if type(w_s[f"J{one_row}"].value) is not str else 0  # 锗

    s_v = {"Cd": cd_v, "Hg": hg_v, "As": as_v, "Pb": pb_v, "Cr": cr_v, "Se": se_v, "Ge": ge_v}
    try:
        yplb = w_s[f"E{one_row}"].value  # 样品类型
        cd_mv = cache_dict[yplb]["Cd"]  # 镉值
        hg_mv = cache_dict[yplb]["Hg"]  # 汞值
        as_mv = cache_dict[yplb]["As"]  # 铅值
        pb_mv = cache_dict[yplb]["Pb"]  # 砷值
        cr_mv = cache_dict[yplb]["Cr"]  # 铬值
        s_mv = {"Cd": cd_mv, "Hg": hg_mv, "As": as_mv, "Pb": pb_mv, "Cr": cr_mv}
        level = 0
        #  判断
        rows_ = 13
        for index_, one_cell in enumerate(["Cd", "Hg", "As", "Pb", "Cr"]):
            if s_v[one_cell] <= s_mv[one_cell]:
                level = 1
            elif s_v[one_cell] <= 2 * s_mv[one_cell]:
                level = 2
            else:
                level = 3
            w_s.cell(one_row, rows_).value = level
            w_s.cell(one_row, rows_ + 1).value = "-"
            w_s.cell(one_row, rows_ + 2).value = "-"
            rows_ += 1
    except:
        rows_ = 13
        for index_, one_cell in enumerate(["Cd", "Hg", "As", "Pb", "Cr"]):
            w_s.cell(one_row, rows_).value = "缺少评判标准"
            w_s.cell(one_row, rows_ + 1).value = "-"
            w_s.cell(one_row, rows_ + 2).value = "-"
            rows_ += 1

    # 搜集等级数据、包名、值等  变量名 level_data_dict   质控数据变量名  zk_list

    for one_data in zk_dict:
        if w_s.cell(one_row, 4).value == one_data:
            level_data_dict[one_data] = [w_s.cell(one_row, 13).value,
                                         w_s.cell(one_row, 14).value,
                                         w_s.cell(one_row, 15).value,
                                         w_s.cell(one_row, 16).value,
                                         w_s.cell(one_row, 17).value,
                                         w_s.cell(one_row, 18).value,
                                         w_s.cell(one_row, 19).value,
                                         w_s.cell(one_row, 6).value,
                                         w_s.cell(one_row, 7).value,
                                         w_s.cell(one_row, 8).value,
                                         w_s.cell(one_row, 9).value,
                                         w_s.cell(one_row, 10).value,
                                         w_s.cell(one_row, 11).value,
                                         w_s.cell(one_row, 12).value,
                                         w_s.cell(one_row, 1).value,
                                         w_s.cell(one_row, 20).value,
                                         w_s.cell(one_row, 4).value,
                                         zk_dict[w_s.cell(one_row, 4).value]]
            zk_list.append(level_data_dict[one_data])
bao_list = []
for one_element in level_data_dict:
    bao_list.append(level_data_dict[one_element][14])
bao_xunhuan_list = list(set(bao_list))
px_s = w_b.create_sheet("平行样质控")
bz_s = w_b.create_sheet("标准物质质控")
# 写入该表抬头
px_title = ["项目", "cd_level", "hg_level", "as_level", "pb_level", "cr_level", "se_level", "ge_level", "cd_value",
            "hg_value", "as_value", "pb_value", "cr_value", "se_value", "ge_value", "包名", "原始编码", "二次编码", "样品类型"]
for _number, one_i in enumerate(px_title):
    px_s.cell(1, _number + 1).value = one_i
    bz_s.cell(1, _number + 1).value = one_i
bz_r = 2  # 标准物质行控制
px_r = 2  # 平行样行控制
for one_bao in bao_xunhuan_list:  # 按包名循环值
    rr = 0
    for one_list in zk_list:  # 在数据列表中寻找
        if one_list[14] == one_bao:  # 数据列表中找到指定包名数据
            if one_list[17].strip() != "标准物质":  # 标准物质数据写入及判断
                rr += 1
                for px_number, px_row in enumerate(one_list):
                    px_s.cell(px_r, px_number + 2).value = px_row
                if rr == 2:
                    px_r += 3
                    px_s.cell(px_r - 2, 1).value = "一致性及相对偏差"
                    # 相对偏差对比  收集两列的数值  upv_list  downv_list
                    upv_list = [px_s.cell(px_r - 4, _upv).value for _upv in range(9, 16)]
                    downv_list = [px_s.cell(px_r - 3, _downv).value for _downv in range(9, 16)]
                    for upv_number, onev_up in enumerate(upv_list):
                        try:
                            pc_v = abs(onev_up - downv_list[upv_number]) / (onev_up + downv_list[upv_number]) * 100
                            px_s.cell(px_r - 2, upv_number + 9).value = pc_v
                        except:
                            px_s.cell(px_r - 2, upv_number + 9).value = "\\"
                    # 一致性对比  收集两列的等级值  up_list  down_list
                    up_list = [px_s.cell(px_r - 4, _up).value for _up in range(2, 9)]
                    down_list = [px_s.cell(px_r - 3, _down).value for _down in range(2, 9)]
                    for up_number, one_up in enumerate(up_list):
                        if down_list[up_number] != one_up:
                            px_s.cell(px_r - 2, up_number + 2).value = "不一致"
                            pc_mv = px_s.cell(px_r - 2, up_number + 9).value
                            yuansu_dict = {2: "Cd", 3: "Hg", 4: "As", 5: "Pb", 6: "Cr", 7: "Se", 8: "Ge"}
                            yuansu = yuansu_dict[up_number + 2]  # 判定为什么元素

                            yuansuzhi = max([px_s.cell(px_r - 3, up_number + 9).value,
                                             px_s.cell(px_r - 4, up_number + 9).value])  # 两个元素的最大值
                            # 与第一个对比
                            pc_xianzhi = 0
                            if yuansuzhi < xdpc_dict[yuansu][0]:
                                pc_xianzhi = xdpc_dict[yuansu][1]
                            elif yuansuzhi < xdpc_dict[yuansu][3]:
                                pc_xianzhi = xdpc_dict[yuansu][4]
                            else:
                                pc_xianzhi = xdpc_dict[yuansu][5]

                            if pc_mv < pc_xianzhi:
                                px_s.cell(px_r - 1, up_number + 9).value = "合格"
                            else:
                                px_s.cell(px_r - 1, up_number + 9).value = "不合格"
                        else:
                            px_s.cell(px_r - 2, up_number + 2).value = "一致"
                            px_s.cell(px_r - 1, up_number + 9).value = "合格"
                    px_s.cell(px_r - 1, 1).value = "是否合格"
                else:
                    px_r += 1
            else:
                # 标准物质编号
                bz_s.cell(bz_r + 1, 1).value = "标准物质含量"
                bz_s.cell(bz_r, 1).value = "原始值"
                bz_bh = one_list[15]
                bz_data = bz_message["农产品"][bz_bh][:7]
                # 写入标准物质检测值
                for bz_number, bz_row in enumerate(one_list):
                    bz_s.cell(bz_r, bz_number + 2).value = bz_row
                bz_s.cell(bz_r + 2, 1).value = "相对偏差值"
                bz_s.cell(bz_r + 3, 1).value = "是否合格"
                # 写入标准物质参考值
                for one_bz, bz_nn in enumerate(bz_data):

                    # 对应元素值
                    yuansu_vv = bz_s.cell(bz_r, one_bz + 9).value if type(bz_s.cell(bz_r, one_bz + 9).value) is not str else 0
                    # 标准物质值
                    bz_s.cell(bz_r + 1, one_bz + 9).value = bz_nn
                    # 判断为何种元素
                    yuansu_bvdict = {2: "Cd", 3: "Hg", 4: "As", 5: "Pb", 6: "Cr", 7: "Se", 8: "Ge"}
                    yuansu_bv = yuansu_bvdict[one_bz + 2]
                    if yuansu_bv in ["Se", "Ge"]:
                        bz_s.cell(bz_r + 3, one_bz + 9).value = "-"
                    else:
                        # 元素值的相对偏差参考值获取
                        pcbv_xianzhi = 0
                        if yuansu_vv < xdpc_dict[yuansu_bv][0]:
                            pcbv_xianzhi = xdpc_dict[yuansu_bv][1]
                        elif yuansu_vv < xdpc_dict[yuansu_bv][3]:
                            pcbv_xianzhi = xdpc_dict[yuansu_bv][4]
                        else:
                            pcbv_xianzhi = xdpc_dict[yuansu_bv][5]
                        # 判断使用何种方法做计算
                        # 第一种含 ± 号值
                        if "±" in str(bz_nn):
                            left_v = float(str(bz_nn)[:str(bz_nn).find("±")])
                            right_v = float(str(bz_nn)[str(bz_nn).find("±") + 1:])
                            up_bv, down_bv = left_v - right_v, left_v + right_v
                            # 元素值与前值的相对偏差
                            bz_xdpc = abs(yuansu_vv - left_v) / (yuansu_vv + left_v) * 100
                            bz_s.cell(bz_r + 2, one_bz + 9).value = bz_xdpc
                            # 判断方法，判定是否在范围值内,或判定与前值的相对偏差是否合理
                            if ((yuansu_vv <= up_bv) and (yuansu_vv >= down_bv)) or (bz_xdpc < pcbv_xianzhi):
                                bz_s.cell(bz_r + 3, one_bz + 9).value = "合格"
                            else:
                                bz_s.cell(bz_r + 3, one_bz + 9).value = "不合格"

                        # 第二中含（）值
                        if "-" in str(bz_nn):
                            # 参考值
                            bz_bv = float(str(bz_nn).strip("-"))
                            # 相对偏差值
                            bz_xdpcer = abs(yuansu_vv - bz_bv) / (yuansu_vv + bz_bv) * 100
                            bz_s.cell(bz_r + 2, one_bz+9).value = bz_xdpcer
                            if bz_xdpcer < pcbv_xianzhi:
                                bz_s.cell(bz_r + 3, one_bz + 9).value = "合格"
                            else:
                                bz_s.cell(bz_r + 3, one_bz + 9).value = "不合格"
                bz_r += 4

w_b.save(os.path.join(fdir_path, "从江县等级判定结果20211128.xlsx"))
