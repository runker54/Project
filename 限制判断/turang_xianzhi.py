#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Date:2021-11-16
# Author:Runker54
# ----------------------
import openpyxl
data_path = r"C:\Users\65680\Desktop\总结报告表格\项目区土壤PH统计表.xlsx"
# 标准字典
select_dict = {"水田": {5: {"Cd": 0.3, "Hg": 0.5, "As": 30, "Pb": 80, "Cr": 250},
                      6: {"Cd": 0.4, "Hg": 0.5, "As": 30, "Pb": 100, "Cr": 250},
                      7: {"Cd": 0.6, "Hg": 0.6, "As": 25, "Pb": 140, "Cr": 300},
                      8: {"Cd": 0.8, "Hg": 1.0, "As": 20, "Pb": 240, "Cr": 350}},
               "其它": {5: {"Cd": 0.3, "Hg": 1.3, "As": 40, "Pb": 70, "Cr": 150},
                      6: {"Cd": 0.3, "Hg": 1.8, "As": 40, "Pb": 90, "Cr": 150},
                      7: {"Cd": 0.3, "Hg": 2.4, "As": 30, "Pb": 120, "Cr": 200},
                      8: {"Cd": 0.6, "Hg": 3.4, "As": 25, "Pb": 170, "Cr": 250}}}

# 限值字典
max_dict = {5: {"Cd": 1.5, "Hg": 2.0, "As": 200, "Pb": 400, "Cr": 800},
            6: {"Cd": 2.0, "Hg": 2.5, "As": 150, "Pb": 500, "Cr": 850},
            7: {"Cd": 3.0, "Hg": 4.0, "As": 120, "Pb": 700, "Cr": 1000},
            8: {"Cd": 4.0, "Hg": 6.0, "As": 100, "Pb": 1000, "Cr": 1300}}

#  读取Excel表格
work_book = openpyxl.load_workbook(data_path)
work_sheet = work_book['Sheet2']
rows = work_sheet.max_row
cols = work_sheet.max_column
# 写结果表头
work_sheet.cell(1, 12).value = "Cd_level"
work_sheet.cell(1, 13).value = "Hg_level"
work_sheet.cell(1, 14).value = "As_level"
work_sheet.cell(1, 15).value = "Pb_level"
work_sheet.cell(1, 16).value = "Cr_level"
#  读取数据并判断

for one_row in range(2, rows+1):
    #  确定字典查询值
    gD = work_sheet.cell(one_row, 3).value  # 耕地类型
    pH = work_sheet.cell(one_row, 4).value  # pH
    try:
        if pH <= 5.5:
            key = 5
        elif pH <= 6.5:
            key = 6
        elif pH <= 7.5:
            key = 7
        else:
            key = 8
    except:
        key = 5
    cD = work_sheet.cell(one_row, 5).value  # 镉
    hG = work_sheet.cell(one_row, 6).value  # 汞
    aS = work_sheet.cell(one_row, 7).value  # 砷
    pB = work_sheet.cell(one_row, 8).value  # 铅
    cR = work_sheet.cell(one_row, 9).value  # 铬
    select_value_Cd = select_dict[gD][key]["Cd"]
    select_value_Hg = select_dict[gD][key]["Hg"]
    select_value_As = select_dict[gD][key]["As"]
    select_value_Pb = select_dict[gD][key]["Pb"]
    select_value_Cr = select_dict[gD][key]["Cr"]
    max_value_Cd = max_dict[key]["Cd"]
    max_value_Hg = max_dict[key]["Hg"]
    max_value_As = max_dict[key]["As"]
    max_value_Pb = max_dict[key]["Pb"]
    max_value_Cr = max_dict[key]["Cr"]

    select_value_dict = {"cD": select_value_Cd, "hG": select_value_Hg, "aS": select_value_As, "pB": select_value_Pb,
                         "cR": select_value_Cr}
    max_value_dict = {"cD": max_value_Cd, "hG": max_value_Hg, "aS": max_value_As, "pB": max_value_Pb, "cR": max_value_Cr}
    key_cell_dict = {"cD": cD, "hG": hG, "aS": aS, "pB": pB, "cR": cR}
    for index_, one_cell in enumerate(["cD", "hG", "aS", "pB", "cR"]):
        if key_cell_dict[one_cell] <= select_value_dict[one_cell]:
            level = 1
        elif key_cell_dict[one_cell] <= max_value_dict[one_cell]:
            level = 2
        else:
            level = 3
        work_sheet.cell(one_row, 12+index_).value = level

work_book.save(r"C:\Users\65680\Desktop\播州区土壤数据评价.xlsx")
