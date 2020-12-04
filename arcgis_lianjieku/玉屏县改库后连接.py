# -*- coding: utf-8 -*-#
# -----------------------
# Date:2020-11-28
# -----------------------
import arcpy
import openpyxl
import time
import xlrd

file_data = r'C:\Users\65680\Desktop\玉屏县\T玉屏县台账数据.gdb\评价单元对应农用地图层_无机5综合_52_Project_SD'
up_sheet = r'C:\Users\65680\Desktop\玉屏县.xlsx'

read_work = openpyxl.load_workbook(up_sheet)  # 工作簿
read_sheet = read_work.get_sheet_by_name('玉屏县')  # 工作表
rows = read_sheet.max_row  # 行
cols = read_sheet.max_column  # 列
# 编码检索词典
bm_dict = {}
for one_row in range(2, rows+1):
    dkbm = read_sheet.cell(one_row, 1).value
    bm_dict[dkbm] = one_row
print(bm_dict)
with arcpy.da.UpdateCursor(file_data, ['标识字段', '玉米19年', '玉米20年', '水稻19年', '水稻20年',
                                       '主要作物19年', '主要作物20年', '周边环境',
                                       '玉米19年面积', '玉米20年面积', '水稻19年面积', '水稻20年面积', 'CQCS']) as cursor:
    # 计数
    count = 0  # 更新数量
    error = 0  # 报错数量
    for one_change in cursor:
        if one_change[12] == '':
            try:

                row_temp = bm_dict[one_change[0]]  # 得到对应编码所在的行数
                one_change[12] = read_sheet.cell(row_temp, 2).value  # 采取措施
                one_change[1] = read_sheet.cell(row_temp, 3).value  # 玉米19年
                one_change[2] = read_sheet.cell(row_temp, 4).value  # 玉米20年
                one_change[3] = read_sheet.cell(row_temp, 5).value  # 水稻19年
                one_change[4] = read_sheet.cell(row_temp, 6).value  # 水稻20年
                one_change[5] = read_sheet.cell(row_temp, 7).value    # 主要作物19年
                one_change[6] = read_sheet.cell(row_temp, 8).value  # 主要作物20年
                cursor.updateRow(one_change)
                count += 1
                # print(f"{one_change[0]}已更新")
            except:
                print(f"{one_change[0]}不在表格内，无法更新")
                error += 1
        else:
            pass
    print(f"成功{count}个，失败{error}个")