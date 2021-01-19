#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Date:2020-12-10
# Author:Runker54
# -----------------------
import xlrd
import openpyxl

# 表格路径
data_path = r'C:\Users\65680\Desktop\毕节市.xlsx'
file_path = '限量值.txt'
text_stream = open(file_path, encoding='utf-8')
new_message_list = []
search_dict = {}
for one in text_stream.readlines():
    x = one.strip('\t\n')
    one_row = x.split('\t')
    new_message_list.append(one_row)
title_list = new_message_list[0]
for one_dict in new_message_list[1:]:
    search_dict[one_dict[0]] = dict(zip(title_list[1:], one_dict[1:]))

# 各元素查询字典
cx_dict = {'水稻': {'Cd': '0.2', 'Hg': '0.02', 'As': '0.2', 'Pb': '0.2', 'Cr': '1'},
           '玉米': {'Cd': '0.1', 'Hg': '0.02', 'As': '0.5', 'Pb': '0.2', 'Cr': '1'},
           '辣椒': {'Cd': '0.05', 'Hg': '0.01', 'As': '0.5', 'Pb': '0.1', 'Cr': '0.5'},
           '红薯': {'Cd': '0.1', 'Hg': '0.01', 'As': '0.5', 'Pb': '0.2', 'Cr': '0.5'},
           '南瓜': {'Cd': '0.1', 'Hg': '0.01', 'As': '0.5', 'Pb': '0.1', 'Cr': '0.5'},
           '茄子': {'Cd': '0.05', 'Hg': '0.01', 'As': '0.5', 'Pb': '0.1', 'Cr': '0.5'},
           '番薯': {'Cd': '0.1', 'Hg': '0.01', 'As': '0.5', 'Pb': '0.2', 'Cr': '0.5'},
           '百香果': {'Cd': '0.05', 'Hg': '0.01', 'As': '0.5', 'Pb': '0.2', 'Cr': '0.5'},
           '佛手瓜': {'Cd': '0.05', 'Hg': '0.01', 'As': '0.5', 'Pb': '0.1', 'Cr': '0.5'},
           '高粱': {'Cd': '0.1', 'Hg': '0.02', 'As': '0.5', 'Pb': '0.2', 'Cr': '1'},
           '生姜': {'Cd': '0.1', 'Hg': '0.01', 'As': '0.5', 'Pb': '3', 'Cr': '0.5'},
           '小米': {'Cd': '0.1', 'Hg': '0.02', 'As': '0.5', 'Pb': '0.2', 'Cr': '1'},
           '四季豆': {'Cd': '0.1', 'Hg': '0.01', 'As': '0.5', 'Pb': '0.2', 'Cr': '0.5'}}
work = openpyxl.load_workbook(data_path)
sheet = work[work.sheetnames[0]]
rows = sheet.max_row  # 获取行数
cols = sheet.max_column  # 获取列数


# 判断等级函数
def calc_level(set_v, god_v):
    """set_v:测量值，god_v:限量值，依据指定规则返回元素等级。"""
    level = 0
    try:
        set_v = float(set_v)
        if set_v <= god_v:
            level = 1
        if god_v < set_v <= 2 * god_v:
            level = 2
        if set_v > 2 * god_v:
            level = 3
    except ValueError:
        if set_v == 'ND':
            level = 1
        if set_v == '无数据':
            level = '无数据'
    return level


# 循环判断写入
for one_row in range(2, rows + 1):
    zw_lx = sheet.cell(one_row, 12).value.strip(' ')
    # 得到各元素测量值
    cd_ = sheet.cell(one_row, 14).value
    hg_ = sheet.cell(one_row, 15).value
    as_ = sheet.cell(one_row, 16).value
    pb_ = sheet.cell(one_row, 17).value
    cr_ = sheet.cell(one_row, 18).value
    # 得到各元素限制
    cd_v = cx_dict[zw_lx]['Cd']
    hg_v = cx_dict[zw_lx]['Hg']
    as_v = cx_dict[zw_lx]['As']
    pb_v = cx_dict[zw_lx]['Pb']
    cr_v = cx_dict[zw_lx]['Cr']
    # 判断并返回值写入
    sheet.cell(one_row, 19, calc_level(cd_, cd_v))
    sheet.cell(one_row, 20, calc_level(hg_, hg_v))
    sheet.cell(one_row, 21, calc_level(as_, as_v))
    sheet.cell(one_row, 22, calc_level(pb_, pb_v))
    sheet.cell(one_row, 23, calc_level(cr_, cr_v))
    sheet.cell(one_row, 24, max(calc_level(cd_, cd_v),
                                calc_level(hg_, hg_v),
                                calc_level(as_, as_v),
                                calc_level(pb_, pb_v),
                                calc_level(cr_, cr_v)))

# 判断后表格输出路径
work.save(r'C:\Users\65680\Desktop\毕节市2.xlsx')
