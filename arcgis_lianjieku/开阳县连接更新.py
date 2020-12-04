# -*- coding: utf-8 -*-#
# -------------------------
# Author:       Runker54
# Date:         2020-11-17
# -------------------------
import arcpy
import openpyxl
import time

file_data = r'F:\T开阳县台账数据.gdb\T无措施数据_确权'
up_sheet = r'C:\Users\65680\Desktop\开阳县农作物调查表.xlsx'

read_work = openpyxl.load_workbook(up_sheet)  # 工作簿
read_sheet = read_work.get_sheet_by_name('name')  # 工作表
rows = read_sheet.max_row  # 行
cols = read_sheet.max_column  # 列
# 编码检索词典
bm_dict = {}
for one_row in range(1, rows):
    dkbm = read_sheet.cell(one_row, 2).value
    bm_dict[dkbm] = one_row
print(bm_dict)
# time.sleep(100)
with arcpy.da.UpdateCursor(file_data, ['地块编码', '玉米19年', '玉米20年', '水稻19年', '水稻20年',
                                       '主要作物19年', '主要作物20年', '周边环境',
                                       '玉米19年面积', '玉米20年面积', '水稻19年面积', '水稻20年面积']) as cursor:
    # 计数
    count = 0  # 更新数量
    error = 0  # 报错数量
    for one_change in cursor:
        try:
            row_temp = bm_dict[one_change[0]]  # 得到对应编码所在的行数
            one_change[1] = read_sheet.cell(row_temp, 8).value  # 玉米19年
            one_change[2] = read_sheet.cell(row_temp, 9).value  # 玉米20年
            one_change[3] = read_sheet.cell(row_temp, 10).value  # 水稻19年
            one_change[4] = read_sheet.cell(row_temp, 11).value  # 水稻20年
            # one_change[5] = read_sheet.cell(row_temp,).value    # 主要作物19年
            one_change[6] = str(read_sheet.cell(row_temp, 12).value) + ' ' + \
                            str(read_sheet.cell(row_temp, 13).value) + ' ' + \
                            str(read_sheet.cell(row_temp, 14).value) + ' ' + \
                            str(read_sheet.cell(row_temp, 15).value)  # 主要作物20年
            one_change[7] = '无'  # 周边环境
            cursor.updateRow(one_change)
            count += 1
            print(f"{one_change[0]}已更新")
        except:
            print(f"{one_change[0]}不在表格内，无法更新")
            error += 1
    print(f"成功{count}个，失败{error}个")
