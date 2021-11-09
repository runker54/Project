#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Date:2021-10-08
# Author:Runker54
# -----------------------
import os
import re
import time
import openpyxl
import random
# list_t = ['惠水县', '罗甸县', '从江县', '兴仁市', '紫云县', '石阡县', '江口县', '思南县', '玉屏县', '大方县', '黔西市', '纳雍县', '赫章县', '织金县', '金沙县',
#           '开阳县', '播州区']
# for x in list_t:
#     patha = r"C:\Users\65680\Desktop\2021年样品检测数据统计.xlsx"
#     work_book = openpyxl.load_workbook(patha)
#     sheet_list = work_book.sheetnames
#     for i in sheet_list:
#         ws = work_book[i]
#         if ws.title != x:
#             work_book.remove_sheet(work_book.get_sheet_by_name(i))
#     work_book.save(r"C:\Users\65680\Desktop\test\2021年{}样品检测数据统计.xlsx".format(x))

change_path = r'C:\Users\65680\Desktop\导出表.xlsx'
data_path = r'C:\Users\65680\Desktop\补充表.xlsx'

data_work = openpyxl.load_workbook(data_path)
data_sheet = data_work.active
data_row = data_sheet.max_row
data_col = data_sheet.max_column
xzc_list = {}
for xzc in ['腊全村', '干团村', '宰门村', '友团村', '腊水村']:
    xzc_ll = []
    for one_row in range(2, data_row + 1):
        row_1_value = data_sheet.cell(one_row, 1).value
        row_2_value = data_sheet.cell(one_row, 2).value
        if row_1_value == xzc:
            xzc_ll.append(row_2_value)
    xzc_ll = list(set(xzc_ll))
    xzc_list[xzc] = xzc_ll
chage_work = openpyxl.load_workbook(change_path)
chage_sheet = chage_work.active
chage_row = chage_sheet.max_row
chage_col = chage_sheet.max_column
for chage_one_row in range(2, chage_row + 1):
    chage_one_row_1 = chage_sheet.cell(chage_one_row, 2).value

    chage_one_row_2 = chage_sheet.cell(chage_one_row, 3).value
    if chage_one_row_2 not in xzc_list[chage_one_row_1]:
        chage_sheet.cell(chage_one_row, 12).value = random.choice(xzc_list[chage_one_row_1])
    else:
        chage_sheet.cell(chage_one_row, 12).value = chage_one_row_2

chage_work.save(r'C:\Users\65680\Desktop\导出表.xlsx')