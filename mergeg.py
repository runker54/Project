#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Date:2021-09-15
# Author:Runker54
# -----------------------
import openpyxl
import os
import time

sheet_path = r'C:\Users\65680\Desktop\新建文件夹'
chrdir_list = os.listdir(sheet_path)
write_book = openpyxl.Workbook()
write_sheet = write_book.create_sheet()
n = 1
for one_file in chrdir_list:
    if one_file.endswith("xlsx"):
        sheet1_path = os.path.join(sheet_path,one_file)
        print(sheet1_path)
        work_book = openpyxl.load_workbook(sheet1_path)
        sheet_list = work_book.worksheets
        for one_sheet in sheet_list:
            row = one_sheet.max_row
            col = one_sheet.max_column
            for one_row in range(5, row + 1):
                for one_col in range(1, col + 1):
                    write_sheet.cell(n, one_col).value = one_sheet.cell(one_row, one_col).value
                n += 1
write_book.save(r'C:\Users\65680\Desktop\从江县入户调查_merge.xlsx')
