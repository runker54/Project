#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Date:2021-11-24
# Author:Runker54
# ----------------------
import MySQLdb
import openpyxl
import time
import os

xls_path = r"C:\Users\65680\Desktop\紫云添加.xlsx"
wb = openpyxl.load_workbook(xls_path)
ws = wb.active
rows = ws.max_row
cols = ws.max_column
host = "rm-wz9clu23kp2cn4lpzbo.mysql.rds.aliyuncs.com"
user = "gdscza"
passwd = "gdsczaGDSCZAcystCYST123!@#"
db = "db_gdscza"
charset = "UTF8"
port = 33006
db = MySQLdb.connect(host=host, port=port, user=user, passwd=passwd, db=db, charset=charset)
for one_row in range(2, rows + 1):
    point_id = ws["AI%s" % one_row].value  # 对应ID
    point_code = ws["A%s" % one_row].value  # 田块编码
    househould_name = ws["B%s" % one_row].value  # 农户姓名
    investigate_target = ws["C%s" % one_row].value  # 调查对象
    investigate_target_phone = ws["D%s" % one_row].value  # 调查对象联系电话
    investigate_user = ws["E%s" % one_row].value  # 调查人
    investigate_user_phone = ws["F%s" % one_row].value  # 调查人联系电话
    villagegroup_name = ws["I%s" % one_row].value  # 组别
    field_type = ws["M%s" % one_row].value  # 土地类型
    org_way = '1'  # 实施方式
    history_yield = ws["N%s" % one_row].value  # 往年产量
    has_monitor = '否'  # 是否设置监测点位
    plant_model = '其它'  # 种植模式
    plant_model_other = '玉米'  # 作物
    around_has_pollute = '否'  # 周边是否有污染源
    last_crop_type = ws["T%s" % one_row].value  # 2020作物名称
    last_variety = ws["U%s" % one_row].value  # 2020作物品种
    last_water_source = "降雨"  # 2020灌溉水来源
    crop_type = ws["X%s" % one_row].value  # 2021作物名称
    variety = ws["Y%s" % one_row].value  # 2021作物品种
    water_source = "降雨"  # 2021灌溉水来源
    complex_manure_name = ws["AB%s" % one_row].value  # 基肥复合肥名称
    complex_manure_ratio = ws["AC%s" % one_row].value  # 基肥复合肥比例
    complex_manure_usage = ws["AD%s" % one_row].value  # 基肥复合肥用量
    complex_manure_name1 = ws["AE%s" % one_row].value  # 追肥复合肥名称
    complex_manure_ratio1 = ws["AF%s" % one_row].value  # 追肥复合肥比例
    complex_manure_usage1 = ws["AG%s" % one_row].value  # 追肥复合肥用量
    name = ws["AJ%s" % one_row].value  # 农药名称
    pusage = ws["AK%s" % one_row].value  # 用量
    usage_age = ws["AL%s" % one_row].value  # 时期
    variety_url = '/files/2021/紫云县入户调查添加/种子/%s.png' % variety
    other_url = '/files/2021/紫云县入户调查添加/农药/%s.jpg' % name
    areabold_url = '/files/2021/紫云县入户调查添加/玉米照片/玉米照片/%s.jpg' % point_code
    investigatetarget_url = '/files/2021/紫云县入户调查添加/农户照片/%s.jpg' % househould_name
    manure_url = '/files/2021/紫云县入户调查添加/复合肥/%s.png' % complex_manure_name
    currsor = db.cursor()
    # 更改基本信息表
    sql = "UPDATE db_gdscza.t_household_survey " \
          "SET db_gdscza.t_household_survey.household_name = '%s', " \
          "db_gdscza.t_household_survey.investigate_target = '%s'," \
          "db_gdscza.t_household_survey.investigate_target_phone = '%s'," \
          "db_gdscza.t_household_survey.investigate_user = '%s'," \
          "db_gdscza.t_household_survey.investigate_user_phone = '%s'," \
          "db_gdscza.t_household_survey.village_group_name = '%s'," \
          "db_gdscza.t_household_survey.field_type = '%s'," \
          "db_gdscza.t_household_survey.org_way = '%s'," \
          "db_gdscza.t_household_survey.history_yield = '%s'," \
          "db_gdscza.t_household_survey.has_monitor = '%s'," \
          "db_gdscza.t_household_survey.plant_model = '%s'," \
          "db_gdscza.t_household_survey.plant_model_other = '%s'," \
          "db_gdscza.t_household_survey.around_has_pollute = '%s'," \
          "db_gdscza.t_household_survey.last_crop_type = '%s'," \
          "db_gdscza.t_household_survey.last_variety = '%s'," \
          "db_gdscza.t_household_survey.last_water_source = '%s'," \
          "db_gdscza.t_household_survey.crop_type = '%s'," \
          "db_gdscza.t_household_survey.variety = '%s'," \
          "db_gdscza.t_household_survey.water_source = '%s'" \
          "WHERE db_gdscza.t_household_survey.point_code = '%s'" % (
              househould_name, investigate_target, investigate_target_phone, investigate_user, investigate_user_phone,
              villagegroup_name, field_type, org_way, history_yield, has_monitor, plant_model, plant_model_other,
              around_has_pollute, last_crop_type, last_variety, last_water_source, crop_type, variety, water_source,
              point_code)
    # print(point_code)
    # 更改肥料表基肥
    sql_jifei = "UPDATE db_gdscza.t_household_survey_manure " \
                "SET db_gdscza.t_household_survey_manure.complex_manure_name = '%s', " \
                "db_gdscza.t_household_survey_manure.complex_manure_ratio = '%s'," \
                "db_gdscza.t_household_survey_manure.complex_manure_usage = '%s'" \
                "WHERE db_gdscza.t_household_survey_manure.point_id = '%s' " \
                "AND db_gdscza.t_household_survey_manure.manure_type = '1'" % (
                    complex_manure_name, complex_manure_ratio, complex_manure_usage, point_id)
    # 更改肥料表追肥
    sql_zuifei = "UPDATE db_gdscza.t_household_survey_manure " \
                 "SET db_gdscza.t_household_survey_manure.complex_manure_name = '%s', " \
                 "db_gdscza.t_household_survey_manure.complex_manure_ratio = '%s'," \
                 "db_gdscza.t_household_survey_manure.complex_manure_usage = '%s'" \
                 "WHERE db_gdscza.t_household_survey_manure.point_id = '%s' " \
                 "AND db_gdscza.t_household_survey_manure.manure_type = '2'" % (
                     complex_manure_name1, complex_manure_ratio1, complex_manure_usage1, point_id)
    # 检查ID
    sql3 = "SELECT object_id " \
           "FROM db_gdscza.sys_file " \
           "WHERE object_id = '%s'" % point_id
    # 插入新行
    sql_insert_j = "INSERT INTO t_household_survey_manure(id, point_id, manure_type) values ('%s1','%s','%s')" % (
        point_code, point_id, '1')
    sql_insert_z = "INSERT INTO t_household_survey_manure(id, point_id, manure_type) values ('%s2','%s','%s')" % (
        point_code, point_id, '2')
    # 更改年份
    sql_chage_year = "UPDATE db_gdscza.t_household_survey_point " \
                     "SET db_gdscza.t_household_survey_point.year = '2021' " \
                     "WHERE db_gdscza.t_household_survey_point.code = '%s'" % point_code
    # 农药表格
    sql_nongyao = "UPDATE db_gdscza.t_household_survey_pesticide " \
                  "SET db_gdscza.t_household_survey_pesticide.name = '%s', " \
                  "db_gdscza.t_household_survey_pesticide.pusage = '%s瓶'," \
                  "db_gdscza.t_household_survey_pesticide.usage_age = '%s'" \
                  "WHERE db_gdscza.t_household_survey_pesticide.point_id = '%s' " % (
                      name, pusage, usage_age, point_id)
    # 插入农药新行
    sql_new_nongyao = "INSERT INTO t_household_survey_pesticide(id, point_id) values ('%s_R','%s')" % (
        point_code, point_id)

    # 插入措施新行
    sql_new_cuoshi_9 = "INSERT INTO t_household_survey_measures(id, point_id,measures_type_code,measures_type_name,point_code) values ('%s_1R','%s','governTechnology','VIP综合/治理技术','%s')" % (
        point_code, point_id, point_code)
    sql_new_cuoshi_7 = "INSERT INTO t_household_survey_measures(id, point_id,measures_type_code,measures_type_name,point_code) values ('%s_2R','%s','waterControl','水分调控','%s')" % (
        point_code, point_id, point_code)
    sql_new_cuoshi_4 = "INSERT INTO t_household_survey_measures(id, point_id,measures_type_code,measures_type_name,point_code) values ('%s_3R','%s','varietyChange','品种调整','%s')" % (
        point_code, point_id, point_code)
    sql_new_cuoshi_3 = "INSERT INTO t_household_survey_measures(id, point_id,measures_type_code,measures_type_name,point_code) values ('%s_4R','%s','leafControl','叶面调控','%s')" % (
        point_code, point_id, point_code)
    sql_new_cuoshi_5 = "INSERT INTO t_household_survey_measures(id, point_id,measures_type_code,measures_type_name,point_code) values ('%s_5R','%s','limeAdjust','石灰调节','%s')" % (
        point_code, point_id, point_code)
    sql_new_cuoshi_1 = "INSERT INTO t_household_survey_measures(id, point_id,measures_type_code,measures_type_name,point_code) values ('%s_6R','%s','deepPloughing','深翻耕','%s')" % (
        point_code, point_id, point_code)
    sql_new_cuoshi_2 = "INSERT INTO t_household_survey_measures(id, point_id,measures_type_code,measures_type_name,point_code) values ('%s_7R','%s','optimizeManure','优化施肥','%s')" % (
        point_code, point_id, point_code)
    sql_new_cuoshi_10 = "INSERT INTO t_household_survey_measures(id, point_id,measures_type_code,measures_type_name,point_code) values ('%s_8R','%s','otherMeasures','其他措施','%s')" % (
        point_code, point_id, point_code)
    sql_new_cuoshi_8 = "INSERT INTO t_household_survey_measures(id, point_id,measures_type_code,measures_type_name,point_code) values ('%s_9R','%s','plantingStructChange','种植结构调整','%s')" % (
        point_code, point_id, point_code)
    sql_new_cuoshi_6 = "INSERT INTO t_household_survey_measures(id, point_id,measures_type_code,measures_type_name,point_code) values ('%s_10R','%s','inSituPassivation','原位钝化','%s')" % (
        point_code, point_id, point_code)
    # 措施计算
    sql_calc_cuoshi = "UPDATE db_gdscza.t_household_survey_measures " \
                      "SET db_gdscza.t_household_survey_measures.has_cheched = '1'" \
                      "WHERE db_gdscza.t_household_survey_measures.point_id = '%s' " \
                      "AND db_gdscza.t_household_survey_measures.measures_type_code = 'varietyChange'" % point_id
    # 插入地址行
    sql_new_dir_variety = "INSERT INTO sys_file(id, object_id, business_type) values ('%s_1R','%s','variety')" % (
        point_code, point_id)
    sql_new_dir_other = "INSERT INTO sys_file(id, object_id, business_type) values ('%s_2R','%s','other')" % (
        point_code, point_id)
    sql_new_dir_manure = "INSERT INTO sys_file(id, object_id, business_type) values ('%s_3R','%s','manure')" % (
        point_code, point_id)
    sql_new_dir_investigateTarget = "INSERT INTO sys_file(id, object_id, business_type) values ('%s_4R','%s','investigateTarget')" % (
        point_code, point_id)
    sql_new_dir_areablod = "INSERT INTO sys_file(id, object_id, business_type) values ('%s_5R','%s','areablod')" % (
        point_code, point_id)
    # 填写地址
    sql_calc_url_1 = "UPDATE db_gdscza.sys_file " \
                     "SET db_gdscza.sys_file.url = '%s'" \
                     "WHERE db_gdscza.sys_file.object_id = '%s' " \
                     "AND db_gdscza.sys_file.business_type = 'variety'" % (variety_url, point_id)
    sql_calc_url_2 = "UPDATE db_gdscza.sys_file " \
                     "SET db_gdscza.sys_file.url = '%s'" \
                     "WHERE db_gdscza.sys_file.object_id = '%s' " \
                     "AND db_gdscza.sys_file.business_type = 'manure'" % (manure_url, point_id)
    sql_calc_url_3 = "UPDATE db_gdscza.sys_file " \
                     "SET db_gdscza.sys_file.url = '%s'" \
                     "WHERE db_gdscza.sys_file.object_id = '%s' " \
                     "AND db_gdscza.sys_file.business_type = 'investigateTarget'" % (investigatetarget_url, point_id)
    sql_calc_url_4 = "UPDATE db_gdscza.sys_file " \
                     "SET db_gdscza.sys_file.url = '%s'" \
                     "WHERE db_gdscza.sys_file.object_id = '%s' " \
                     "AND db_gdscza.sys_file.business_type = 'areablod'" % (areabold_url, point_id)
    sql_calc_url_5 = "UPDATE db_gdscza.sys_file " \
                     "SET db_gdscza.sys_file.url = '%s'" \
                     "WHERE db_gdscza.sys_file.object_id = '%s' " \
                     "AND db_gdscza.sys_file.business_type = 'other'" % (other_url, point_id)

    # currsor.execute(sql)
    # currsor.execute(sql2)
    # currsor.execute(sql3)
    # currsor.execute(sql3)
    # data = currsor.fetchall()
    # cahe_chae = len(data)
    # print()
    # if cahe_chae == 0:
    #     currsor.execute(sql_calc_url_1)
    #     currsor.execute(sql_calc_url_2)
    #     currsor.execute(sql_calc_url_3)
    #     currsor.execute(sql_calc_url_4)
    #     currsor.execute(sql_calc_url_5)
    #     print("插入新行")
    # else:
    #     # currsor.execute(sql_calc_url_1)
    #     # currsor.execute(sql_calc_url_2)
    #     currsor.execute(sql_calc_url_3)
    #     # currsor.execute(sql_calc_url_4)
    #     # currsor.execute(sql_calc_url_5)
    # currsor.execute(sql_calc_url_3)
    # 检查状态
    sql_check = "UPDATE db_gdscza.t_household_survey_point " \
                "SET db_gdscza.t_household_survey_point.check_status = '1'," \
                "db_gdscza.t_household_survey_point.check_level = '8'," \
                "db_gdscza.t_household_survey_point.check_node = '【省级单位】已审核【通过】'" \
                "WHERE db_gdscza.t_household_survey_point.id = '%s' " % point_id
    currsor.execute(sql_check)
    currsor.close()
    print(point_code, point_id)

db.commit()
db.close()
print(1)
