#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Date:2021-12-09
# Author:Runker54
# ----------------------
import MySQLdb
import openpyxl
import time
import os

xls_path = r"C:\Users\65680\Desktop\change2.xlsx"
wb = openpyxl.load_workbook(xls_path)
ws = wb["Sheet2"]
rows = ws.max_row
host = "rm-wz9clu23kp2cn4lpzbo.mysql.rds.aliyuncs.com"
user = "gdscza"
passwd = "gdsczaGDSCZAcystCYST123!@#"
dbases = "db_gdscza"
charset = "UTF8"
port = 33006
db = MySQLdb.connect(host=host, port=port, user=user, passwd=passwd, db=dbases, charset=charset)
ncol = 0  # 计数
for one_row in range(2, rows + 1):
    currsor = db.cursor()
    point_id = ws["A%s" % one_row].value  # 对应ID
    point_code = ws["B%s" % one_row].value  # 田块编码
    househould_name = ws["O%s" % one_row].value  # 农户姓名
    investigate_target = ws["O%s" % one_row].value  # 调查对象
    investigate_target_phone = ws["P%s" % one_row].value  # 调查对象联系电话
    investigate_user = ws["M%s" % one_row].value  # 调查人
    investigate_user_phone = ws["N%s" % one_row].value  # 调查人联系电话
    villagegroup_name = ws["H%s" % one_row].value  # 组别
    field_type = ws["Q%s" % one_row].value  # 土地类型
    org_way = '1'  # 实施方式
    history_yield = ws["V%s" % one_row].value  # 往年产量
    has_monitor = ws["AB%s" % one_row].value  # 是否设置监测点位
    plant_model = ws["W%s" % one_row].value  # 种植模式
    plant_model_other = ws["X%s" % one_row].value  # 作物
    around_has_pollute = ws["R%s" % one_row].value  # 周边是否有污染源
    last_crop_type = ws["S%s" % one_row].value  # 2020作物名称
    last_variety = ws["T%s" % one_row].value  # 2020作物品种
    last_water_source = ws["U%s" % one_row].value  # 2020灌溉水来源
    crop_type = ws["Y%s" % one_row].value  # 2021作物名称
    variety = ws["Z%s" % one_row].value  # 2021作物品种
    water_source = ws["AA%s" % one_row].value  # 2021灌溉水来源
    complex_manure_name = ws["AG%s" % one_row].value  # 基肥复合肥名称
    complex_manure_ratio = ws["AH%s" % one_row].value  # 基肥复合肥比例
    complex_manure_usage = ws["AI%s" % one_row].value  # 基肥复合肥用量
    complex_manure_name1 = ws["AQ%s" % one_row].value  # 追肥复合肥名称
    complex_manure_ratio1 = ws["AR%s" % one_row].value  # 追肥复合肥比例
    complex_manure_usage1 = ws["AS%s" % one_row].value  # 追肥复合肥用量
    name = ws["AC%s" % one_row].value  # 农药名称
    pusage = str(ws["AD%s" % one_row].value) + "ml"  # 用量
    usage_age = ws["AE%s" % one_row].value  # 时期
    cuoshi_name = ws["AF%s" % one_row].value  # 措施名称
    # 更新基础信息表
    sql_survey = "UPDATE db_gdscza.t_household_survey " \
                 "SET db_gdscza.t_household_survey.village_group_name = '%s'," \
                 "db_gdscza.t_household_survey.field_type = '%s'," \
                 "db_gdscza.t_household_survey.org_way = '%s'," \
                 "db_gdscza.t_household_survey.history_yield = '%s'," \
                 "db_gdscza.t_household_survey.has_monitor = '%s'," \
                 "db_gdscza.t_household_survey.plant_model = '%s'," \
                 "db_gdscza.t_household_survey.plant_model_other = '%s'," \
                 "db_gdscza.t_household_survey.around_has_pollute = '%s'," \
                 "db_gdscza.t_household_survey.last_crop_type = '%s'," \
                 "db_gdscza.t_household_survey.last_variety = '%s'," \
                 "db_gdscza.t_household_survey.last_water_source = '%s'," \
                 "db_gdscza.t_household_survey.crop_type = '%s'," \
                 "db_gdscza.t_household_survey.variety = '%s'," \
                 "db_gdscza.t_household_survey.water_source = '%s'" \
                 "WHERE db_gdscza.t_household_survey.point_code = '%s'" % (
                     villagegroup_name, field_type, org_way, history_yield, has_monitor, plant_model, plant_model_other,
                     around_has_pollute, last_crop_type, last_variety, last_water_source, crop_type, variety,
                     water_source,
                     point_code)
    # 更新基肥肥料情况表
    sql_survey_manure_1 = "UPDATE db_gdscza.t_household_survey_manure " \
                          "SET db_gdscza.t_household_survey_manure.complex_manure_name = '%s', " \
                          "db_gdscza.t_household_survey_manure.complex_manure_ratio = '%s'," \
                          "db_gdscza.t_household_survey_manure.complex_manure_usage = '%s'" \
                          "WHERE db_gdscza.t_household_survey_manure.point_id = '%s' " \
                          "AND db_gdscza.t_household_survey_manure.manure_type = '1'" % (
                              complex_manure_name, complex_manure_ratio, complex_manure_usage, point_id)
    # 更新追肥肥料情况表
    sql_survey_manure_2 = "UPDATE db_gdscza.t_household_survey_manure " \
                          "SET db_gdscza.t_household_survey_manure.complex_manure_name = '%s', " \
                          "db_gdscza.t_household_survey_manure.complex_manure_ratio = '%s'," \
                          "db_gdscza.t_household_survey_manure.complex_manure_usage = '%s'" \
                          "WHERE db_gdscza.t_household_survey_manure.point_id = '%s' " \
                          "AND db_gdscza.t_household_survey_manure.manure_type = '2'" % (
                              complex_manure_name1, complex_manure_ratio1, complex_manure_usage1, point_id)
    # 更新农药情况表
    sql_survey_pesticide = "UPDATE db_gdscza.t_household_survey_pesticide " \
                           "SET db_gdscza.t_household_survey_pesticide.name = '%s', " \
                           "db_gdscza.t_household_survey_pesticide.pusage = '%s'," \
                           "db_gdscza.t_household_survey_pesticide.usage_age = '%s'" \
                           "WHERE db_gdscza.t_household_survey_pesticide.point_id = '%s' " % (
                               name, pusage, usage_age, point_id)
    # 更新措施填写情况

    cuoshi_dict = {"品种调整": 0, "叶面调控": 1, "种植结构调整": 2, "深翻耕": 3, "优化施肥": 4, "原位钝化": 5, "VIP综合/治理技术": 6, "石灰调节": 7,
                   "其它措施": 8, "水分调控": 9}
    sql_survey_meaures_default = "UPDATE db_gdscza.t_household_survey_measures " \
                                 "SET db_gdscza.t_household_survey_measures.has_cheched = '0'" \
                                 "WHERE db_gdscza.t_household_survey_measures.point_id = '%s' " % point_id
    sql_survey_meaures_variety = "UPDATE db_gdscza.t_household_survey_measures " \
                                 "SET db_gdscza.t_household_survey_measures.has_cheched = '1'" \
                                 "WHERE db_gdscza.t_household_survey_measures.point_id = '%s' " \
                                 "AND db_gdscza.t_household_survey_measures.measures_type_code = 'varietyChange'" % point_id
    sql_survey_meaures_leafControl = "UPDATE db_gdscza.t_household_survey_measures " \
                                     "SET db_gdscza.t_household_survey_measures.has_cheched = '1'" \
                                     "WHERE db_gdscza.t_household_survey_measures.point_id = '%s' " \
                                     "AND db_gdscza.t_household_survey_measures.measures_type_code = 'leafControl'" % point_id
    sql_survey_meaures_optimizeManure = "UPDATE db_gdscza.t_household_survey_measures " \
                                     "SET db_gdscza.t_household_survey_measures.has_cheched = '1'" \
                                     "WHERE db_gdscza.t_household_survey_measures.point_id = '%s' " \
                                     "AND db_gdscza.t_household_survey_measures.measures_type_code = 'optimizeManure'" % point_id
    sql_survey_meaures_waterControl = "UPDATE db_gdscza.t_household_survey_measures " \
                                     "SET db_gdscza.t_household_survey_measures.has_cheched = '1'" \
                                     "WHERE db_gdscza.t_household_survey_measures.point_id = '%s' " \
                                     "AND db_gdscza.t_household_survey_measures.measures_type_code = 'waterControl'" % point_id

    currsor.execute(sql_survey)  # 执行更新基础信息表
    print("基础信息表更新完成")
    currsor.execute(sql_survey_manure_1)  # 执行更新基肥肥料表
    print("基肥肥料表更新完成")
    currsor.execute(sql_survey_manure_2)  # 执行更新追肥肥料表
    print("追肥肥料表更新完成")
    currsor.execute(sql_survey_pesticide)  # 执行更新农药信息表
    print("农药信息表更新完成")
    currsor.execute(sql_survey_meaures_default)  # 执行更新措施信息表
    if str(cuoshi_name).strip() in ["同类作物品种调整", "调整种植低富集作物", "调整种植非食用作物", "品种调整"]:
        currsor.execute(sql_survey_meaures_variety)
    if str(cuoshi_name).strip() == "叶面调控":
        currsor.execute(sql_survey_meaures_leafControl)
    if str(cuoshi_name).strip() == "优化施肥":
        currsor.execute(sql_survey_meaures_optimizeManure)
    if str(cuoshi_name).strip() == "水分调控":
        currsor.execute(sql_survey_meaures_waterControl)
    if str(cuoshi_name).strip() == "品种调整+水分调控":
        currsor.execute(sql_survey_meaures_variety)
        currsor.execute(sql_survey_meaures_waterControl)
    if str(cuoshi_name).strip() == "品种调整+叶面调控":
        currsor.execute(sql_survey_meaures_variety)
        currsor.execute(sql_survey_meaures_leafControl)
    if str(cuoshi_name).strip() == "品种调整+优化施肥":
        currsor.execute(sql_survey_meaures_variety)
        currsor.execute(sql_survey_meaures_optimizeManure)
    if str(cuoshi_name).strip() == "品种调整+水分调控+叶面调控":
        currsor.execute(sql_survey_meaures_variety)
        currsor.execute(sql_survey_meaures_waterControl)
        currsor.execute(sql_survey_meaures_leafControl)
    print("措施信息表更新完成")
    print(point_code)
    ncol += 1
    currsor.close()
    db.commit()
    print(f"已提交{ncol}条记录")
db.close()
