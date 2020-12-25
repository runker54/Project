#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Date:2020-12-18
# Author:Runker54
# -----------------------
import sys
from PyQt5.QtWidgets import QMainWindow, QApplication, QFileDialog
# from PyQt5.QtGui import *
import xlrd
import xlwt
import openpyxl
import os
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QTimer, QDateTime
import random


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(800, 600)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        # 按钮
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)  # 编码转换按钮
        self.pushButton.setGeometry(QtCore.QRect(20, 320, 100, 50))
        self.pushButton.setObjectName("pushButton")
        self.pushButton_2 = QtWidgets.QPushButton(self.centralwidget)  # 等级评价按钮
        self.pushButton_2.setGeometry(QtCore.QRect(20, 390, 100, 50))
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_3 = QtWidgets.QPushButton(self.centralwidget)  # 合并单个book多个sheet按钮
        self.pushButton_3.setGeometry(QtCore.QRect(140, 320, 100, 50))
        self.pushButton_3.setObjectName("pushButton_3")
        self.pushButton_4 = QtWidgets.QPushButton(self.centralwidget)  # 合并单个book多个sheet按钮
        self.pushButton_4.setGeometry(QtCore.QRect(140, 390, 100, 50))
        self.pushButton_4.setObjectName("pushButton_4")
        # 标签
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(20, 40, 72, 15))
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(430, 40, 72, 15))
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.centralwidget)   # 输出详细信息
        self.label_3.setGeometry(QtCore.QRect(430, 300, 101, 16))  # 输出详细信息
        self.label_3.setObjectName("label_3")  # 输出详细信息
        self.label_4 = QtWidgets.QLabel(self.centralwidget)   # at信息
        self.label_4.setGeometry(QtCore.QRect(20, 560, 250, 20))  # at信息  x, y, length, height
        self.label_4.setObjectName("label_4")  # at信息
        # 文本框
        self.textEdit = QtWidgets.QTextEdit(self.centralwidget)
        self.textEdit.setGeometry(QtCore.QRect(20, 70, 341, 211))
        self.textEdit.setObjectName("textEdit")
        self.textEdit_2 = QtWidgets.QTextEdit(self.centralwidget)
        self.textEdit_2.setGeometry(QtCore.QRect(430, 70, 341, 211))
        self.textEdit_2.setObjectName("textEdit_2")
        self.textEdit_3 = QtWidgets.QTextEdit(self.centralwidget)
        self.textEdit_3.setGeometry(QtCore.QRect(430, 330, 341, 211))
        self.textEdit_3.setObjectName("textEdit_3")

        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 800, 26))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.retranslateUi(MainWindow)
        # 触发信号
        self.pushButton.clicked.connect(MainWindow.id_change)
        self.pushButton_2.clicked.connect(MainWindow.level_calc)
        self.pushButton_3.clicked.connect(MainWindow.merge_sheet)
        self.pushButton_4.clicked.connect(MainWindow.chose_directory)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))

        self.label.setText(_translate("MainWindow", "源文件："))
        self.label_2.setText(_translate("MainWindow", "目标文件："))
        self.label_3.setText(_translate("MainWindow", "提示信息："))
        self.label_4.setText(_translate("MainWindow", "Runker54"))

        self.pushButton.setText(_translate("MainWindow", "编码转换"))
        self.pushButton_2.setText(_translate("MainWindow", "等级评价"))
        self.pushButton_3.setText(_translate("MainWindow", "单个表格合并"))
        self.pushButton_4.setText(_translate("MainWindow", "多个表格合并"))


class DetailUI(Ui_MainWindow, QMainWindow):
    def __init__(self):
        super(DetailUI, self).__init__()
        self.setupUi(self)
        self.setWindowTitle('样品转码及农产品评价')
        self.textEdit_3.setText('一、编码转换\n\n'
                                '1、表格需为xls类型，且需转换的sheet在第一个，编码位于第一列。\n\n'
                                '二、等级评价\n\n'
                                '1、表格格式为xlsx类型，作物位于第2列，Cd,Hg,As,Pb,Cr元素检测值分别位于第3，4，5，6，7列.\n\n'
                                '2、元素检测值不能为空。\n\n'
                                '三、单个表格合并\n\n'
                                '1、针对一个表格内含有多个子表的情况；传入表格实现子表合并为一个表（xls类型）。\n\n'
                                '四、多个表格合并\n\n'
                                '1、合并同一文件夹下的所有表格包括含有子表的情况（xls类型）')

    def id_change(self):
        try:
            fname, _ = QFileDialog.getOpenFileName(self, 'open file', '/', "Excel files (*.xls *.xlsx)")
            self.textEdit.setText(fname)
            xls_path = fname  # 得到文件路径
            temp_path = os.path.splitext(xls_path)
            new_path = temp_path[0] + "_hash转码" + temp_path[1]
            xls_work_book = xlrd.open_workbook(xls_path)
            xls_sheet = xls_work_book.sheet_by_index(0)
            xls_rows = xls_sheet.nrows
            # xls_cols = xls_sheet.ncols
            r = 0
            bao_number = 1
            new_work_book = xlwt.Workbook('utf-8')
            new_sheet = new_work_book.add_sheet("包%s" % bao_number, cell_overwrite_ok=True)  # 创建新列表
            # 写入标题
            for id_nember, one_title in enumerate(['采样编码', '所属包', '样品类型', '二次编码']):
                new_sheet.write(0, id_nember, one_title)
            zky_number = ''
            zky_id = random.randint(1, 47)
            for one_row in range(1, xls_rows):
                yd_id = str(xls_sheet.row(one_row)[0].value).strip(' ')
                new_sheet.write(one_row - (bao_number - 1) * 47, 0, xls_sheet.row(one_row)[0].value)
                new_sheet.write(one_row - (bao_number - 1) * 47, 1, "包%s" % bao_number)
                new_sheet.write(one_row - (bao_number - 1) * 47, 2, "检测样")
                new_sheet.write(one_row - (bao_number - 1) * 47, 3, str(abs(hash(yd_id))))
                if one_row - (bao_number - 1) * 47 == zky_id:
                    zky_number = str(xls_sheet.row(one_row)[0].value).strip(' ')
                    new_sheet.write(one_row - (bao_number - 1) * 47, 2, "质控样")
                # 写入其它内容
                # for one_col in range(1, xls_cols):
                #     new_sheet.write(one_row-(bao_number-1)*47, one_col + 1, xls_sheet.row(one_row)[one_col].value)
                r += 1
                if r == 47:
                    r = 0
                    bao_number += 1
                    # 写入后三个样品
                    sn_zky = '520000%szksn' % bao_number
                    sj_zky = '520000%szksj' % bao_number
                    zky_number_string = '%sZK' % zky_number
                    for next_id, next_string in enumerate(['质控样', '室内质控样', '室间质控样']):
                        if next_id == 1:
                            next_id = next_id + 48
                            new_sheet.write(next_id, 0, '室内质控样-%s' % (bao_number-1))
                            new_sheet.write(next_id, 1, "包%s" % (bao_number-1))
                            new_sheet.write(next_id, 2, next_string)
                            new_sheet.write(next_id, 3, str(abs(hash(sn_zky))))
                        if next_id == 2:
                            next_id = next_id + 48
                            new_sheet.write(next_id, 0, '室间质控样-%s' % (bao_number-1))
                            new_sheet.write(next_id, 1, "包%s" % (bao_number-1))
                            new_sheet.write(next_id, 2, next_string)
                            new_sheet.write(next_id, 3, str(abs(hash(sj_zky))))
                        if next_id == 0:
                            next_id = next_id + 48
                            new_sheet.write(next_id, 0, zky_number_string)
                            new_sheet.write(next_id, 1, "包%s" % (bao_number-1))
                            new_sheet.write(next_id, 2, next_string)
                            new_sheet.write(next_id, 3, str(abs(hash(zky_number_string))))
                    new_sheet = new_work_book.add_sheet("包%s" % bao_number, cell_overwrite_ok=True)  # 创建新列表
                    for id_nember, one_title in enumerate(['采样编码', '所属包', '样品类型', '二次编码']):
                        new_sheet.write(0, id_nember, one_title)
                    zky_id = random.randint(1, 47)
            new_work_book.save(new_path)
            self.textEdit.setText("所选择文件:%s转码完成！" % fname)
            self.textEdit_2.setText("输出文件路径:%s" % new_path)
        except:
            self.textEdit.setText("所选择文件类型有误或表格格式不正确！")

    def level_calc(self):
        try:
            fname, _ = QFileDialog.getOpenFileName(self, 'open file', '/', "Excel files (*.xls *.xlsx)")
            self.textEdit.setText(fname)
            xls_path = fname  # 得到文件路径
            print(xls_path)
            temp_path = os.path.splitext(xls_path)
            new_path = temp_path[0] + "_等级计算" + temp_path[1]
            cx_dict = {'水稻': {'Cd': '0.2', 'Hg': '0.02', 'As': '0.2', 'Pb': '0.2', 'Cr': '1'},
                       '玉米': {'Cd': '0.1', 'Hg': '0.02', 'As': '0.5', 'Pb': '0.2', 'Cr': '1'},
                       '辣椒': {'Cd': '0.05', 'Hg': '0.01', 'As': '0.5', 'Pb': '0.1', 'Cr': '0.5'},
                       '红薯': {'Cd': '0.1', 'Hg': '0.01', 'As': '0.5', 'Pb': '0.2', 'Cr': '0.5'},
                       '南瓜': {'Cd': '0.1', 'Hg': '0.01', 'As': '0.5', 'Pb': '0.1', 'Cr': '0.5'},
                       '茄子': {'Cd': '0.05', 'Hg': '0.01', 'As': '0.5', 'Pb': '0.1', 'Cr': '0.5'},
                       '番薯': {'Cd': '0.1', 'Hg': '0.01', 'As': '0.5', 'Pb': '0.2', 'Cr': '0.5'},
                       '百香果': {'Cd': '0.05', 'Hg': '0.01', 'As': '0.5', 'Pb': '0.2', 'Cr': '0.5'},
                       '佛手瓜': {'Cd': '0.05', 'Hg': '0.01', 'As': '0.5', 'Pb': '0.1', 'Cr': '0.5'},
                       '高粱': {'Cd': '0.1', 'Hg': '0.02', 'As': '0.5', 'Pb': '0.2', 'Cr': '1'},
                       '生姜': {'Cd': '0.1', 'Hg': '0.01', 'As': '0.5', 'Pb': '3', 'Cr': '0.5'},
                       '小麦': {'Cd': '0.1', 'Hg': '0.02', 'As': '0.5', 'Pb': '0.2', 'Cr': '1'},
                       '四季豆': {'Cd': '0.1', 'Hg': '0.01', 'As': '0.5', 'Pb': '0.2', 'Cr': '0.5'}}  # 各作物查询字典
            work = openpyxl.load_workbook(xls_path)
            sheet = work[work.sheetnames[0]]
            rows = sheet.max_row  # 获取行数
            # 判断等级函数

            def calc_level(set_v, god_v):
                """set_v:测量值，god_v:限量值，依据指定规则返回元素等级。"""
                level = ''
                god_v = float(god_v)
                try:
                    set_v = float(set_v)
                    if set_v <= god_v:
                        level = '1'
                    if god_v < set_v <= 2 * god_v:
                        level = '2'
                    if set_v > 2 * god_v:
                        level = '3'
                except ValueError:
                    if set_v == 'ND':
                        level = '1'
                    if set_v == '无数据':
                        level = '无数据'
                    else:
                        level = ''
                return level

            # 循环判断写入

            for one_row in range(2, rows + 1):
                zw_lx = sheet.cell(one_row, 2).value.strip(' ')
                # 得到各元素测量值
                cd_ = sheet.cell(one_row, 3).value
                hg_ = sheet.cell(one_row, 4).value
                as_ = sheet.cell(one_row, 5).value
                pb_ = sheet.cell(one_row, 6).value
                cr_ = sheet.cell(one_row, 7).value
                # 得到各元素限制
                cd_v = cx_dict[zw_lx]['Cd']
                hg_v = cx_dict[zw_lx]['Hg']
                as_v = cx_dict[zw_lx]['As']
                pb_v = cx_dict[zw_lx]['Pb']
                cr_v = cx_dict[zw_lx]['Cr']
                # 判断并返回值写入
                sheet.cell(one_row, 8, calc_level(cd_, cd_v))
                sheet.cell(one_row, 9, calc_level(hg_, hg_v))
                sheet.cell(one_row, 10, calc_level(as_, as_v))
                sheet.cell(one_row, 11, calc_level(pb_, pb_v))
                sheet.cell(one_row, 12, calc_level(cr_, cr_v))
                sheet.cell(one_row, 13, max(calc_level(cd_, cd_v),
                                            calc_level(hg_, hg_v),
                                            calc_level(as_, as_v),
                                            calc_level(pb_, pb_v),
                                            calc_level(cr_, cr_v)))
            work.save(new_path)
            self.textEdit.setText("所选择文件:%s等级计算完成！" % fname)
            self.textEdit_2.setText("输出文件路径:%s" % new_path)
        except:
            self.textEdit_3.setText("错误原因：\n1、数据表内各元素数值存在空值！\n"
                                    "2、限量值字典不包含表格内作物。")

    def merge_sheet(self):
        try:
            fname, _ = QFileDialog.getOpenFileName(self, 'open file', '/', "Excel files (*.xls *.xlsx)")
            self.textEdit.setText(fname)
            xls_path = fname  # 得到文件路径
            print(xls_path)
            temp_path = os.path.splitext(xls_path)
            new_path = temp_path[0] + "_mergesheet" + temp_path[1]
            xls_work = xlrd.open_workbook(xls_path)
            xls_sheet = xls_work.sheets()
            new_work = xlwt.Workbook('utf-8')
            new_sheet = new_work.add_sheet('merge_sheet')
            r = 1  # 开始行
            # 写入标题
            for id_nember, one_title in enumerate(['采样编码', '所属包', '样品类型', '二次编码']):
                new_sheet.write(0, id_nember, one_title)
            for i in range(len(xls_sheet)):
                read_sheet = xls_sheet[i]
                rows = read_sheet.nrows
                cols = read_sheet.ncols
                for one_row in range(1, rows):
                    for one_col in range(cols):
                        new_sheet.write(r, one_col, read_sheet.row(one_row)[one_col].value)
                    r += 1
            new_work.save(new_path)
            self.textEdit.setText("所选择文件:%s合并完成！" % fname)
            self.textEdit_2.setText("输出文件路径:%s" % new_path)
        except:
            self.textEdit.setText("所选择文件类型有误或表格格式不正确！")

    def chose_directory(self):
        try:
            directory = QtWidgets.QFileDialog.getExistingDirectory(None, "选取文件夹", "C:/")  # 起始路径
            self.textEdit.setText(directory)
            xls_list = []
            self.calc_list(directory, xls_list)
            save_book = xlwt.Workbook("utf-8")
            new_sheet = save_book.add_sheet("merge_sheet")
            r = 1  # 开始行
            for one_xls in xls_list:
                read_xls = xlrd.open_workbook(one_xls)  # 打开一个xls文件
                read_sheets = read_xls.sheets()  # 得到xls文件的所有sheet列表
                for read_sheet in read_sheets:
                    rows = read_sheet.nrows
                    cols = read_sheet.ncols
                    for one_row in range(1, rows):
                        for one_col in range(cols):
                            new_sheet.write(r, one_col, read_sheet.row(one_row)[one_col].value)
                        r += 1
            title_sheet = xlrd.open_workbook(xls_list[0]).sheets()[0]
            for id_, title_ in enumerate([title_sheet.row(0)[zero_col].value for zero_col in range(title_sheet.ncols)]):
                new_sheet.write(0, id_, title_)
            new_path = os.path.join(directory, "merge_sheet.xls")
            save_book.save(new_path)
            self.textEdit.setText("所选择文件夹:%s下的xls文件合并完成！" % directory)
            self.textEdit_2.setText("合并后的文件为:%s" % new_path)
        except:
            self.textEdit.setText("所选择文件夹不包含xls文件")

    def calc_list(self, directory, xls_list):
        for roots, dirs, files in os.walk(directory):
            for one_file in files:
                if one_file.endswith("xls"):
                    one_file = os.path.join(roots, one_file)
                    xls_list.append(one_file)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    ex = DetailUI()
    ex.show()
    sys.exit(app.exec_())
