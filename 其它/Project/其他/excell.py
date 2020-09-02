#!/usr/bin/env python
# -*- coding:utf-8 -*-
# Time:2020/3/25
# import xlwt
# import xlrd
# import os
#
# sheet = xlrd.open_workbook('C:\Users\runke\Desktop\exceltext\CD.xlsx')
# table = sheet.sheet_by_index(0)i
import openpyxl as op
import os

import xlwt

wb = op.load_workbook('C:\\Users\\runke\Desktop\\test.xlsx')
popresuLt = {}
resultList = []
listTem = []
ws = wb.active
for row in range(2, ws.max_row + 1):
    xz = ws['A' + str(row)].value
    lb = ws['B' + str(row)].value
    mj = ws['C' + str(row)].value
    popresuLt.setdefault(xz, {})
    popresuLt[xz].setdefault(lb, {"sl": 0, "mj": 0})
    popresuLt[xz][lb]["sl"] += 1
    popresuLt[xz][lb]["mj"] += mj
for xz, lb_info in popresuLt.items():
    for lb, info in lb_info.items():
        listTem.append(xz)
        listTem.append(lb)
        listTem.append(info['sl'])
        listTem.append(info['mj'])
    resultList.append(listTem)
print(resultList)
listTem = []
numberRow = 0
countSheet = wb.create_chartsheet(index=1, title="统计结果")
for key, value in popresuLt.items():
    # print(key)
    # print(value)
    numberRow += len(value)
    # print(numberRow)
    # countSheet(1,1).value = "乡镇"
    # countSheet(2,1).value = "类别"
    # countSheet(3,1).value = "数量"
    # countSheet(4,1).value = "面积"
    for row in range(2, numberRow + 1):
        countSheet['A' + str(row)] = resultList[row - 2][0]
        countSheet['B' + str(row)] = resultList[row - 2][1]
        countSheet['C' + str(row)] = resultList[row - 2][2]
        countSheet['D' + str(row)] = resultList[row - 2][3]
wb.save("CD.xlsx")
