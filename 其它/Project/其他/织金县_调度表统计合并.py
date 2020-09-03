# -*- coding: utf-8 -*-#
# -------------------------------------------------------------------------------
# Author:       A
# Date:         2020-09-03
# -------------------------------------------------------------------------------
import xlrd
from openpyxl import Workbook
import os
document_path = r"C:\Users\65680\Desktop\新建文件夹"
# 遍历文档路径下的所有xls文件
new_work_book = Workbook()
new_work_sheet = new_work_book.create_sheet("data")
r = 1  # 开始行
for roots, dirs, files in os.walk(document_path):
    for file in files:
        if file.endswith('xls'):
            one_file_path = os.path.join(roots, file)
            one_work_book = xlrd.open_workbook(one_file_path)
            one_work_sheet = one_work_book.sheet_by_index(0)
            rows = one_work_sheet.nrows
            columns = one_work_sheet.ncols
            # print(f"行：{rows}\n列：{columns}")
            for one_row in range(1, rows):
                for columns_ in range(columns):
                    cell_value = one_work_sheet.row(one_row)[columns_].value
                    print(cell_value)
                    new_work_sheet.cell(r, columns_+1).value = cell_value
                r += 1
new_work_book.save(r"C:\Users\65680\Desktop\安全利用类.xlsx")
