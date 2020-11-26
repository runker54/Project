# coing:utf-8
import xlwt
from openpyxl import load_workbook
import os
import time

data_path = r"C:\Users\ols\Desktop\表格处理20201104\在产"
new_workbook = xlwt.Workbook('utf-8')
excel_list = []
for roots, dirs, files in os.walk(data_path):  # 获取所有表格的路径
    for one_file in files:
        if one_file.endswith('xlsx'):
            one_file_adress = os.path.join(roots, one_file)
            excel_list.append(one_file_adress)
sheet_name = []
work_book_getname = load_workbook(excel_list[0])
sheet_name = work_book_getname.get_sheet_names()
for one_sheet in sheet_name:  # 按表名循环
    new_sheet = new_workbook.add_sheet(one_sheet, cell_overwrite_ok=False)
    r = 1  # 开始行
    i = 0  # 判断是否写入表头
    for one_excel in excel_list:
        work_book = load_workbook(one_excel)
        work_book_sheet = work_book.get_sheet_by_name(one_sheet)
        rows = work_book_sheet.max_row
        columns = work_book_sheet.max_column
        title_list = []
        for one_col in range(1, columns + 1):  # 得到表头
            title_one = work_book_sheet.cell(1, one_col).value
            title_list.append(title_one)
        if i == 0:
            for _cols, one_title in enumerate(title_list):  # 写入表头
                new_sheet.write(0, _cols, one_title)
        else:
            pass
        for one_rows in range(2, rows + 1):
            for one_cols in range(1, columns + 1):
                cell_value = work_book_sheet.cell(one_rows, one_cols).value
                new_sheet.write(r, one_cols - 1, cell_value)
            r += 1
        i += 1
new_workbook.save(r'C:\Users\ols\Desktop\合并保存\在产企业.xls')
