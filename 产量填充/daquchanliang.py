#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Date:2021-11-01
# Author:Runker54
# ----------------------
# !/usr/bin/env python
# -*- coding: utf-8 -*-
# Date:2021-10-25
# Author:Runker54
# ---------------------
import openpyxl
import random
import os
import numpy as np
from openpyxl.drawing.image import Image
import time

dir_path = r"C:\Users\65680\Desktop\大区测产"
image_path = r"C:\Users\65680\Desktop\湄潭县图片"
# 信息字典
# 折干率

source_xls = r"C:\Users\65680\Desktop\湄潭县.xlsx"
source_work_book = openpyxl.load_workbook(source_xls)
ws = source_work_book['大区']
rows = ws.max_row
for one_row in range(2, rows + 1):
    moban_xls = r"C:\Users\65680\Desktop\大区.xlsx"
    source_mb_work_book = openpyxl.load_workbook(moban_xls)
    ws_mb = source_mb_work_book.active
    # 折干率
    zgl = random.randint(79, 83)
    # 向模板表填充内容
    # 田块编号
    tkbh = ws[f"D{one_row}"].value
    ws_mb["B4"].value = tkbh
    # 测产人员
    ccry = ws[f"E{one_row}"].value
    ws_mb["E4"].value = ccry
    # 种植品种
    zzpz = ws[f"F{one_row}"].value
    ws_mb["G4"].value = zzpz
    # 测产类型
    cclx = "全田实测"
    ws_mb["K4"].value = cclx
    # 小区面积
    xqmj = str(round(ws[f"G{one_row}"].value, 2)) + "亩"
    ws_mb["N4"].value = xqmj
    # 移栽方式
    yzfs = ws[f"H{one_row}"].value
    ws_mb["B5"].value = yzfs
    # 实施措施
    sscs = ws[f"J{one_row}"].value
    ws_mb["E5"].value = sscs
    # 试验小区处理名称
    syxqclmc = ws[f"I{one_row}"].value
    ws_mb["H5"].value = syxqclmc
    # 测产日期
    ccrq = ws[f"K{one_row}"].value
    ws_mb["M5"].value = ccrq
    # 出田鲜重
    ctxz = round(ws[f"L{one_row}"].value, 1)
    ws_mb["C7"].value = ctxz
    ws_mb["M7"].value = ctxz
    # 折干率
    a_z = zgl - (random.randrange(1, 3) * 0.1)-(random.randrange(1, 9) * 0.01)
    b_z = zgl - (random.randrange(1, 3) * 0.1)-(random.randrange(1, 9) * 0.01)
    c_z = zgl - (random.randrange(1, 3) * 0.1)-(random.randrange(1, 9) * 0.01)
    pj_z = (a_z + b_z + c_z) / 3
    ws[f"M{one_row}"].value = pj_z
    ws_mb["C8"].value = a_z
    ws_mb["E8"].value = b_z
    ws_mb["G8"].value = c_z
    ws_mb["L8"].value = pj_z
    # 折亩产
    zmc = ctxz * pj_z * 0.01 / ws[f"G{one_row}"].value
    ws_mb["C28"].value = str(round(zmc, 1))
    ws[f"N{one_row}"].value = zmc
    # 验收图形
    image_key = ws[f"D{one_row}"].value
    img = Image(os.path.join(image_path, f"无标题_{image_key}0.png"))
    img.width = 722
    img.height = 601
    ws_mb.add_image(img, "B9")
    # 地块面积
    ws_mb["M11"].value = xqmj
    source_mb_work_book.save(os.path.join(dir_path, f"{tkbh}-{ws.cell(one_row, 1).value}.xlsx"))

source_work_book.save(r"C:\Users\65680\Desktop\湄潭测产信息_大区.xlsx")
