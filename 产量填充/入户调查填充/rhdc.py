#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Date:2021-11-03
# Author:Runker54
# ----------------------
import openpyxl
import os
import time
from openpyxl.styles import Font

out_path = r"C:\Users\65680\Desktop\从江县集中推进区入户调查表"
wb_path = r"C:\Users\65680\Desktop\从江县集中推进区赋值信息20211104.xlsx"
wb_sfpz = r"C:\Users\65680\Desktop\品种调整_水分调控入户调查表.xlsx"
wb_work = openpyxl.load_workbook(wb_path)
ws = wb_work["入户调查信息汇总"]
rows = ws.max_row
ws_mb_work = openpyxl.load_workbook(wb_sfpz)
ws_m = ws_mb_work.active
for one_row in range(2, rows + 1):
    # 调查单位
    dcdw = "贵州雏阳生态环保科技有限公司"
    ws_m["B2"].value = dcdw
    # 调查对象
    dcdx = ws[f"D{one_row}"].value
    ws_m["G2"].value = dcdx
    # 调查对象联系电话
    dcdxdh = ws[f"Y{one_row}"].value
    ws_m["H2"].value = dcdxdh
    # 调查人
    dcr = ws[f"AB{one_row}"].value
    ws_m["C3"].value = dcr
    # 调查人联系电话
    dcrdh = ws[f"AC{one_row}"].value
    ws_m["E3"].value = dcrdh
    # 调查日期
    dcrq = ws[f"Z{one_row}"].value
    ws_m["H3"].value = dcrq
    # 田块编号
    tkbh = ws[f"G{one_row}"].value
    ws_m["C4"].value = tkbh
    # 农户姓名
    nhxm = ws[f"D{one_row}"].value
    ws_m["G4"].value = nhxm
    # 所在乡镇
    xz = ws[f"E{one_row}"].value
    ws_m["C5"].value = xz
    # 所在村组
    cz = ws[f"F{one_row}"].value
    ws_m["G5"].value = cz
    # 田块面积
    mj = ws[f"B{one_row}"].value
    ws_m["C6"].value = mj
    # 中心坐标
    zxzb = str(round(ws[f"H{one_row}"].value, 6)) + "," + str(round(ws[f"I{one_row}"].value, 6))
    ws_m["G6"].value = zxzb
    # 往年产量
    wncl = ws[f"N{one_row}"].value
    ws_m["C8"].value = wncl
    # 地貌类型
    dmlx = ws[f"J{one_row}"].value
    ws_m["F8"].value = dmlx
    # 耕作方式
    gzfs = ws[f"AA{one_row}"].value
    ws_m["I8"].value = gzfs
    # 20农作物类型
    tzw20 = ws[f"L{one_row}"].value
    ws_m["C9"].value = tzw20
    # 20农作物品种
    tzw20pz = ws[f"M{one_row}"].value
    ws_m["H9"].value = tzw20pz
    # 20灌溉水来源
    ggs20 = ws[f"K{one_row}"].value
    ws_m["C10"].value = ggs20
    # 21农作物类型
    tzw21 = ws[f"O{one_row}"].value
    ws_m["C11"].value = tzw21
    # 21农作物品种
    tzw21pz = ws[f"P{one_row}"].value
    ws_m["H11"].value = tzw21pz
    # 21灌溉水来源
    ggs21 = ws[f"K{one_row}"].value
    ws_m["C12"].value = ggs21
    # 21基肥种类
    jf21 = ws[f"R{one_row}"].value
    ws_m["D13"].value = jf21
    # 21基肥用量
    jf21yl = ws[f"S{one_row}"].value
    ws_m["H13"].value = jf21yl
    # 21追肥种类
    zf21 = ws[f"T{one_row}"].value
    ws_m["D15"].value = zf21
    # 21追肥用量
    zf21yl = ws[f"U{one_row}"].value
    ws_m["H15"].value = zf21yl
    # 21农药种类
    ny21 = ws[f"AD{one_row}"].value
    ws_m["C17"].value = ny21
    # 土地类型
    ws_m["C7"].value = "□旱地 ☑水田 □其他"
    # 周边污染源
    ws_m["H7"].value = "是□（  ）否☑"
    # 种植模式水田
    ws_m["C18"].value = "☑稻 + 鱼          □稻 + 虾          □稻 + 鸭         □其他："
    # 种植模式旱地
    ws_m["C19"].value = "□玉米+马铃薯      □玉米+豆类         □其他：   "
    # 措施1
    pd_value = ws[f"V{one_row}"].value
    if pd_value == "水分调控":
        ws_m["C20"].value = "□石灰调节         □优化施肥          □品种调整"
    else:
        ws_m["C20"].value = "□石灰调节         □优化施肥          ☑品种调整"
    # 措施2
    ws_m["C21"].value = "☑水分调控         □叶面调控          □深翻耕"
    # 措施3
    ws_m["C22"].value = "□原位钝化         □种植结构调整      □“VIP”综合/治理技术"
    # 措施4
    ws_m["C23"].value = "□其他措施   "
    # 组织方式
    ws_m["C26"].value = "☑自主实施    □第三方实施    □自主+第三方实施"

    ws_mb_work.save(os.path.join(out_path, str(tkbh) + ".xlsx"))
    print(tkbh)