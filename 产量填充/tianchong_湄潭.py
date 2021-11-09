#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Date:2021-10-25
# Author:Runker54
# ---------------------
import openpyxl
import random
import os
import numpy as np
from openpyxl.drawing.image import Image

dir_path = r"C:\Users\65680\Desktop\小区测产"
image_path = r"C:\Users\65680\Desktop\image"
# 信息字典
# 株高
zhugao = {"宜香优2115": [110, 132], "花香7号": [102, 118], "大粒香": [117, 126], "宜香4245": [104, 115], "中优177": [105, 124],
          "双优505": [102, 121], "宜香优2168": [114, 122], "渝香203": [115, 129], "玉针香": [114, 125], "蓉7优2115": [103, 115]}

# 穗长
suichang = {"宜香优2115": [22, 39], "花香7号": [24, 267], "大粒香": [21, 27], "宜香4245": [22, 29], "中优177": [21, 28],
            "双优505": [21, 30], "宜香优2168": [22, 28], "渝香203": [22, 31], "玉针香": [22, 30], "蓉7优2115": [25, 31]}

# 穗粒数
suilishu = {"宜香优2115": [185, 210], "花香7号": [175, 203], "大粒香": [166, 212], "宜香4245": [171, 208],
            "中优177": [179, 219], "双优505": [181, 222], "宜香优2168": [176, 206], "渝香203": [158, 198],
            "玉针香": [178, 208], "蓉7优2115": [199, 225]}
# 千粒重
qianlizhong = {"宜香优2115": 29.2, "花香7号": 28.1, "大粒香": 30.6, "宜香4245": 30.0, "中优177": 28.9, "双优505": 27.6,
               "宜香优2168": 29.3, "渝香203": 30.2, "玉针香": 28.1, "蓉7优2115": 29.6}

# 穗数
yousuishu = {"宜香优2115": [16, 22], "花香7号": [20, 23], "大粒香": [21, 24], "宜香4245": [22, 28], "中优177": [21, 27],
             "双优505": [21, 27], "宜香优2168": [22, 26], "渝香203": [24, 30], "玉针香": [20, 28], "蓉7优2115": [19, 26]}
source_xls = r"C:\Users\65680\Desktop\湄潭县.xlsx"
source_work_book = openpyxl.load_workbook(source_xls)
ws = source_work_book['小区']
rows = ws.max_row
for one_row in range(2, rows + 1):
    if ws[f"G{one_row}"].value == 0.03:
        moban_xls = r"C:\Users\65680\Desktop\4米小区.xlsx"
    elif ws[f"G{one_row}"].value == 0.0375:
        moban_xls = r"C:\Users\65680\Desktop\5米小区.xlsx"
    else:
        moban_xls = r"C:\Users\65680\Desktop\8米小区.xlsx"
    source_mb_work_book = openpyxl.load_workbook(moban_xls)
    ws_mb = source_mb_work_book.active
    # 有穗数
    suishu_list = [random.randint(int(yousuishu[ws.cell(one_row, 6).value][0]),
                                  int(yousuishu[ws.cell(one_row, 6).value][1])) for i in range(10)]
    sum_suishu = sum(suishu_list)
    ws[f"O{one_row}"].value = sum_suishu
    # 亩有效穗数
    mu_youxiao = ws[f"N{one_row}"].value * sum_suishu / 10
    ws[f"P{one_row}"].value = mu_youxiao
    # 穗长
    sc_list = [random.randint(int(suichang[ws.cell(one_row, 6).value][0]),
                              int(suichang[ws.cell(one_row, 6).value][1])) for i in range(10)]
    sum_sc = sum(sc_list)
    ws[f"Q{one_row}"].value = sum_sc
    # 株高
    zg_list = [random.randint(int(zhugao[ws.cell(one_row, 6).value][0]),
                              int(zhugao[ws.cell(one_row, 6).value][1])) for i in range(10)]
    sum_zg = sum(zg_list)
    ws[f"R{one_row}"].value = sum_zg
    # 穗粒数
    sl_list = [random.randint(int(suilishu[ws.cell(one_row, 6).value][0]),
                              int(suilishu[ws.cell(one_row, 6).value][1])) for i in range(10)]
    sum_sl = sum(sl_list)
    ws[f"S{one_row}"].value = sum_sl
    # 穗实粒数
    ssl_list = [int(ssl * random.randint(86, 98) * 0.01) for ssl in sl_list]
    ws[f"T{one_row}"].value = sum(ssl_list) / 10
    # 空秕数
    ws[f"U{one_row}"].value = ws[f"S{one_row}"].value - ws[f"T{one_row}"].value
    # 结实率
    ws[f"V{one_row}"].value = ws[f"T{one_row}"].value / ws[f"S{one_row}"].value
    # 千粒重
    qlz = qianlizhong[ws.cell(one_row, 6).value]
    ws[f"W{one_row}"].value = qlz
    # 粒型
    ws[f"X{one_row}"].value = "长粒"
    # 一千克籽粒重
    qkzlz = 1000 * random.randint(75, 82) * 0.01
    ws[f"Y{one_row}"].value = qkzlz
    # 折干率
    zgl = ws[f"Y{one_row}"].value / 1000
    ws[f"Z{one_row}"].value = zgl
    # 折亩产
    zmc = ws[f"C{one_row}"].value / 2 * ws[f"Z{one_row}"].value / ws[f"G{one_row}"].value
    ws[f"AA{one_row}"].value = zmc
    # 理论产量
    llcl = ws[f"P{one_row}"].value * ws[f"T{one_row}"].value * ws[f"W{one_row}"].value * ws[
        f"Z{one_row}"].value * 0.000001
    ws[f"AB{one_row}"].value = llcl
    # 向模板表填充内容
    # 田块编号
    tkbh = ws[f"D{one_row}"].value
    ws_mb["B4"].value = tkbh
    # 测产人员
    ccry = ws[f"E{one_row}"].value
    ws_mb["E4"].value = ccry
    # 种植品种
    zzpz = ws[f"F{one_row}"].value
    ws_mb["G4"].value = zzpz
    # 测产类型
    cclx = "理论+实测"
    ws_mb["K4"].value = cclx
    # 小区面积
    xqmj = str(ws[f"G{one_row}"].value) + "亩"
    ws_mb["N4"].value = xqmj
    # 移栽方式
    yzfs = ws[f"H{one_row}"].value
    ws_mb["B5"].value = yzfs
    # 实施措施
    sscs = ws[f"J{one_row}"].value
    ws_mb["E5"].value = sscs
    # 试验小区处理名称
    syxqclmc = ws[f"I{one_row}"].value
    ws_mb["H5"].value = syxqclmc
    # 测产日期
    ccrq = ws[f"K{one_row}"].value
    ws_mb["M5"].value = ccrq
    # 十行长度 十窝长度
    shcd = str(ws[f"L{one_row}"].value) + "米"
    swcd = str(ws[f"M{one_row}"].value) + "米"
    ws_mb["C7"].value = shcd
    ws_mb["M7"].value = shcd
    ws_mb["C8"].value = swcd
    ws_mb["M8"].value = swcd
    # 亩穴数
    mxs = ws[f"N{one_row}"].value
    ws_mb["C9"].value = mxs
    # 十穴有穗数
    for yss_number, one_yss in enumerate(suishu_list):
        ws_mb.cell(10, yss_number + 3).value = one_yss
    ws_mb["M10"].value = sum_suishu / 10
    # 亩有效穗
    myxs = mu_youxiao
    ws_mb["C11"].value = myxs
    # 穗长
    for sc_number, one_sc in enumerate(sc_list):
        ws_mb.cell(12, sc_number + 3).value = one_sc
    ws_mb["M12"].value = sum_sc / 10
    # 株高
    for zg_number, one_zg in enumerate(zg_list):
        ws_mb.cell(16, zg_number + 3).value = one_zg
    ws_mb["M16"].value = sum_zg / 10
    # 平均穗粒数
    for sl_number, one_sl in enumerate(sl_list):
        ws_mb.cell(20, sl_number + 3).value = one_sl
    ws_mb["M20"].value = sum_sl / 10
    # 平均穗实粒数
    for ssl_number, one_ssl in enumerate(ssl_list):
        ws_mb.cell(21, ssl_number + 3).value = one_ssl
    ws_mb["M21"].value = sum(ssl_list) / 10
    # 平均穗空秕数
    _a = np.array(sl_list)
    _b = np.array(ssl_list)
    skp_list = list(_a - _b)
    for skp_number, one_skp in enumerate(skp_list):
        ws_mb.cell(22, skp_number + 3).value = one_skp
    ws_mb["M22"].value = sum(skp_list) / 10
    # 结实率
    ws_mb["C23"].value = str(round(sum(ssl_list) / sum(sl_list) * 100, 2)) + "%"
    # 千粒重
    ws_mb["C24"].value = qlz
    # 粒型
    ws_mb["B25"].value = "长粒"
    # 验收图形
    # img = Image(os.path.join(image_path, "T16.png"))
    # img.width = 281
    # img.height = 178
    # ws_mb.add_image(img,"B26")
    # 小区穗重
    ws_mb["C32"].value = round(ws[f"C{one_row}"].value / 2, 1)
    # 1千克籽粒晒干重
    ws_mb["C33"].value = qkzlz
    # 折干率
    ws_mb["C34"].value = round(zgl, 2)
    # 折亩产
    ws_mb["C35"].value = round(zmc, 1)
    # 理论产量
    ws_mb["J35"].value = round(llcl, 1)
    source_mb_work_book.save(os.path.join(dir_path, f"{tkbh}-{ws.cell(one_row, 1).value}.xlsx"))

source_work_book.save(r"C:\Users\65680\Desktop\湄潭测产信息_小区.xlsx")
