#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Date:2021-10-25
# Author:Runker54
# ---------------------
import openpyxl
import random
import os
import numpy as np
from openpyxl.drawing.image import Image

dir_path = r"C:\Users\65680\Desktop\小区测产"
image_path = r"C:\Users\65680\Desktop\image"
# 信息字典
# 株高
zhugao = {"宜香优2115": [111, 130], "川优6203": [108, 122], "荃优1606": [111, 121], "湘优109": [100, 111], "兆优5455": [112, 130],
          "香早优2017": [99, 121], "宜香2239": [114, 126], "宜优673": [118, 128], "恒丰优426": [108, 118], "天优华占": [99, 115],
          "Lcd": [98, 118], "成优1479": [109, 123], "花优357": [108, 125], "兆优6377": [117, 126], "香龙优2018": [106, 122],
          "中优808": [125, 142], "宜香4245": [108, 121]}
# 穗长
suichang = {"宜香优2115": [22, 33], "川优6203": [20, 26], "荃优1606": [21, 27], "湘优109": [22, 30], "兆优5455": [20, 28],
            "香早优2017": [21, 32], "宜香2239": [22, 26], "宜优673": [19, 31], "恒丰优426": [22, 30], "天优华占": [25, 33],
            "Lcd": [24, 32], "成优1479": [25, 31], "花优357": [21, 30], "兆优6377": [24, 34], "香龙优2018": [20, 28],
            "中优808": [20, 26], "宜香4245": [23, 30]}
# 穗粒数
suilishu = {"宜香优2115": [165, 261], "川优6203": [155, 263], "荃优1606": [146, 252], "湘优109": [171, 198],
            "兆优5455": [159, 249], "香早优2017": [171, 252], "宜香2239": [166, 226], "宜优673": [118, 228],
            "恒丰优426": [158, 218], "天优华占": [199, 265], "Lcd": [128, 228], "成优1479": [159, 239], "花优357": [138, 225],
            "兆优6377": [117, 246], "香龙优2018": [136, 252], "中优808": [135, 242], "宜香4245": [138, 261]}
# 千粒重
qianlizhong = {"宜香优2115": 29.2, "川优6203": 28.1, "荃优1606": 30.6, "湘优109": 30.0, "兆优5455": 28.9, "香早优2017": 27.6,
               "宜香2239": 29.3, "宜优673": 30.2, "恒丰优426": 28.1, "天优华占": 29.6, "Lcd": 27.7, "成优1479": 30.1, "花优357": 30.5,
               "兆优6377": 28.8, "香龙优2018": 27.9, "中优808": 31.1, "宜香4245": 28.7}

# 穗数
yousuishu = {"宜香优2115": [16, 32], "川优6203": [20, 26], "荃优1606": [21, 27], "湘优109": [22, 30], "兆优5455": [20, 28],
             "香早优2017": [21, 30], "宜香2239": [22, 26], "宜优673": [19, 31], "恒丰优426": [23, 30], "天优华占": [22, 33],
             "Lcd": [14, 28], "成优1479": [15, 31], "花优357": [21, 30], "兆优6377": [24, 34], "香龙优2018": [20, 28],
             "中优808": [20, 26], "宜香4245": [23, 30]}

source_xls = r"C:\Users\65680\Desktop\播州测产信息111.xlsx"
source_work_book = openpyxl.load_workbook(source_xls)
ws = source_work_book['小区测产信息']
rows = ws.max_row
for one_row in range(2, rows + 1):
    if ws[f"G{one_row}"].value == 0.0240:
        moban_xls = r"C:\Users\65680\Desktop\4米小区.xlsx"
    elif ws[f"G{one_row}"].value == 0.0276:
        moban_xls = r"C:\Users\65680\Desktop\5米小区.xlsx"
    else:
        moban_xls = r"C:\Users\65680\Desktop\8米小区.xlsx"
    source_mb_work_book = openpyxl.load_workbook(moban_xls)
    ws_mb = source_mb_work_book.active
    # 有穗数
    suishu_list = [random.randint(int(yousuishu[ws.cell(one_row, 6).value][0]),
                                  int(yousuishu[ws.cell(one_row, 6).value][1])) for i in range(10)]
    sum_suishu = sum(suishu_list)
    ws[f"O{one_row}"].value = sum_suishu
    # 亩有效穗数
    mu_youxiao = ws[f"N{one_row}"].value * sum_suishu / 10
    ws[f"P{one_row}"].value = mu_youxiao
    # 穗长
    sc_list = [random.randint(int(suichang[ws.cell(one_row, 6).value][0]),
                              int(suichang[ws.cell(one_row, 6).value][1])) for i in range(10)]
    sum_sc = sum(sc_list)
    ws[f"Q{one_row}"].value = sum_sc
    # 株高
    zg_list = [random.randint(int(zhugao[ws.cell(one_row, 6).value][0]),
                              int(zhugao[ws.cell(one_row, 6).value][1])) for i in range(10)]
    sum_zg = sum(zg_list)
    ws[f"R{one_row}"].value = sum_zg
    # 穗粒数
    sl_list = [random.randint(int(suilishu[ws.cell(one_row, 6).value][0]),
                              int(suilishu[ws.cell(one_row, 6).value][1])) for i in range(10)]
    sum_sl = sum(sl_list)
    ws[f"S{one_row}"].value = sum_sl
    # 穗实粒数
    ssl_list = [int(ssl * random.randint(86, 98) * 0.01) for ssl in sl_list]
    ws[f"T{one_row}"].value = sum(ssl_list) / 10
    # 空秕数
    ws[f"U{one_row}"].value = ws[f"S{one_row}"].value - ws[f"T{one_row}"].value
    # 结实率
    ws[f"V{one_row}"].value = ws[f"T{one_row}"].value / ws[f"S{one_row}"].value
    # 千粒重
    qlz = qianlizhong[ws.cell(one_row, 6).value]
    ws[f"W{one_row}"].value = qlz
    # 粒型
    ws[f"X{one_row}"].value = "长粒"
    # 一千克籽粒重
    qkzlz = 1000 * random.randint(70, 79) * 0.01
    ws[f"Y{one_row}"].value = qkzlz
    # 折干率
    zgl = ws[f"Y{one_row}"].value / 1000
    ws[f"Z{one_row}"].value = zgl
    # 折亩产
    zmc = ws[f"C{one_row}"].value / 2 * ws[f"Z{one_row}"].value / ws[f"G{one_row}"].value
    ws[f"AA{one_row}"].value = zmc
    # 理论产量
    llcl = ws[f"P{one_row}"].value * ws[f"T{one_row}"].value * ws[f"W{one_row}"].value * ws[
        f"Z{one_row}"].value * 0.000001
    ws[f"AB{one_row}"].value = llcl
    # 向模板表填充内容
    # 田块编号
    tkbh = ws[f"D{one_row}"].value
    ws_mb["B4"].value = tkbh
    # 测产人员
    ccry = ws[f"E{one_row}"].value
    ws_mb["E4"].value = ccry
    # 种植品种
    zzpz = ws[f"F{one_row}"].value
    ws_mb["G4"].value = zzpz
    # 测产类型
    cclx = "理论+实测"
    ws_mb["K4"].value = cclx
    # 小区面积
    xqmj = str(ws[f"G{one_row}"].value) + "亩"
    ws_mb["N4"].value = xqmj
    # 移栽方式
    yzfs = ws[f"H{one_row}"].value
    ws_mb["B5"].value = yzfs
    # 实施措施
    sscs = ws[f"J{one_row}"].value
    ws_mb["E5"].value = sscs
    # 试验小区处理名称
    syxqclmc = ws[f"I{one_row}"].value
    ws_mb["H5"].value = syxqclmc
    # 测产日期
    ccrq = ws[f"K{one_row}"].value
    ws_mb["M5"].value = ccrq
    # 十行长度 十窝长度
    shcd = str(ws[f"L{one_row}"].value) + "米"
    swcd = str(ws[f"M{one_row}"].value) + "米"
    ws_mb["C7"].value = shcd
    ws_mb["M7"].value = shcd
    ws_mb["C8"].value = swcd
    ws_mb["M8"].value = swcd
    # 亩穴数
    mxs = ws[f"N{one_row}"].value
    ws_mb["C9"].value = mxs
    # 十穴有穗数
    for yss_number, one_yss in enumerate(suishu_list):
        ws_mb.cell(10, yss_number + 3).value = one_yss
    ws_mb["M10"].value = sum_suishu / 10
    # 亩有效穗
    myxs = mu_youxiao
    ws_mb["C11"].value = myxs
    # 穗长
    for sc_number, one_sc in enumerate(sc_list):
        ws_mb.cell(12, sc_number + 3).value = one_sc
    ws_mb["M12"].value = sum_sc / 10
    # 株高
    for zg_number, one_zg in enumerate(zg_list):
        ws_mb.cell(16, zg_number + 3).value = one_zg
    ws_mb["M16"].value = sum_zg / 10
    # 平均穗粒数
    for sl_number, one_sl in enumerate(sl_list):
        ws_mb.cell(20, sl_number + 3).value = one_sl
    ws_mb["M20"].value = sum_sl / 10
    # 平均穗实粒数
    for ssl_number, one_ssl in enumerate(ssl_list):
        ws_mb.cell(21, ssl_number + 3).value = one_ssl
    ws_mb["M21"].value = sum(ssl_list) / 10
    # 平均穗空秕数
    _a = np.array(sl_list)
    _b = np.array(ssl_list)
    skp_list = list(_a - _b)
    for skp_number, one_skp in enumerate(skp_list):
        ws_mb.cell(22, skp_number + 3).value = one_skp
    ws_mb["M22"].value = sum(skp_list) / 10
    # 结实率
    ws_mb["C23"].value = str(round(sum(ssl_list) / sum(sl_list) * 100, 2)) + "%"
    # 千粒重
    ws_mb["C24"].value = qlz
    # 粒型
    ws_mb["B25"].value = "长粒"
    # 验收图形
    # img = Image(os.path.join(image_path, "T16.png"))
    # img.width = 281
    # img.height = 178
    # ws_mb.add_image(img,"B26")
    # 小区穗重
    ws_mb["C32"].value = round(ws[f"C{one_row}"].value/2, 1)
    # 1千克籽粒晒干重
    ws_mb["C33"].value = qkzlz
    # 折干率
    ws_mb["C34"].value = round(zgl, 2)
    # 折亩产
    ws_mb["C35"].value = round(zmc, 1)
    # 理论产量
    ws_mb["J35"].value = round(llcl, 1)
    source_mb_work_book.save(os.path.join(dir_path, f"{tkbh}-{ws.cell(one_row, 1).value}.xlsx"))

source_work_book.save(r"C:\Users\65680\Desktop\播州测产信息_小区.xlsx")

